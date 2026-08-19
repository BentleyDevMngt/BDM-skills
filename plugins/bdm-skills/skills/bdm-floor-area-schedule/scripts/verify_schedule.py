#!/usr/bin/env python3
"""Reconcile the markup PDF against the Form 405 workbook — cell by cell.

Nothing in the schedule is reported until this passes clean.  It reads the
workbook (calculated values), the take-off config, and the markup PDF, and
checks that every figure printed on a drawing matches the workbook cell it
claims to come from.

    python3 verify_schedule.py workbook.xlsx config.json markups.pdf

Exit code 0 = clean, 1 = at least one figure does not tie.
"""
import json, sys

import openpyxl
import pymupdf

TOL = 0.05          # m² — figures are reported to 1 dp


def f(v):
    return f'{v:,.1f}'


def main():
    if len(sys.argv) != 4:
        sys.exit(__doc__)
    wbf, cfgf, pdff = sys.argv[1:]
    wb = openpyxl.load_workbook(wbf, data_only=True)
    cfg = json.load(open(cfgf))
    pages = [p.get_text() for p in pymupdf.open(pdff)]
    allt = '\n'.join(pages)

    fails, checks = [], 0

    def has(page, s, what):
        nonlocal checks
        checks += 1
        t = pages[page - 1] if page else allt
        if s not in t:
            fails.append(f'p{page or "-"}  MISSING {what}:  {s!r}')

    def near(a, b, what):
        nonlocal checks
        checks += 1
        if a is None or b is None or abs(a - b) > TOL:
            fails.append(f'{what}: {a} != {b}')

    ls, nr, ac = wb['Level Schedule'], wb['NSA & Revenue'], wb['Area by Category']

    # index of pages by (level, kind), cover is page 1
    page_of = {(sh['level'], sh['kind']): i for i, sh in enumerate(cfg['sheets'], 2)}

    # ---- level figures: workbook Level Schedule -> config -> cover -> plan sheet
    row = {}
    r = 16
    while ls.cell(r, 1).value and str(ls.cell(r, 1).value).upper() != 'TOTAL':
        row[str(ls.cell(r, 1).value)] = r
        r += 1
    for lv in cfg['levels']:
        n = lv['name']
        if n not in row:
            fails.append(f'level {n!r} is on the markup but not in the workbook Level Schedule')
            continue
        rr = row[n]
        for key, col in (('feca', 2), ('uca', 3), ('gfa', 4), ('gba', 5)):
            near(lv.get(key), ls.cell(rr, col).value, f'{n} {key.upper()} config vs workbook')
        has(1, f(ls.cell(rr, 2).value), f'cover FECA {n}')
        has(1, f(ls.cell(rr, 4).value), f'cover GFA {n}')
        has(1, f(ls.cell(rr, 5).value), f'cover GBA {n}')
        p = page_of.get((n, 'mk'))
        if p:
            has(p, f'FECA {f(ls.cell(rr, 2).value)}', f'{n} sheet FECA')
            has(p, f'UCA {f(ls.cell(rr, 3).value or 0)}', f'{n} sheet UCA')
            has(p, f'GFA {f(ls.cell(rr, 4).value)}', f'{n} sheet GFA')
            has(p, f'GBA {f(ls.cell(rr, 5).value)}', f'{n} sheet GBA')
    tot_r = r
    for col, name in ((2, 'FECA'), (3, 'UCA'), (4, 'GFA'), (5, 'GBA')):
        has(1, f(ls.cell(tot_r, col).value), f'cover total {name}')

    # ---- apartment NSA: workbook NSA & Revenue -> config -> cover -> plan sheet
    cfg_apt = {a['apt']: a for a in cfg['apartments']}
    lvl_tot = {}
    r = 10
    while nr.cell(r, 1).value and str(nr.cell(r, 1).value).upper() != 'TOTAL':
        apt, lv = str(nr.cell(r, 1).value), str(nr.cell(r, 2).value)
        internal, bal, tot = nr.cell(r, 4).value, nr.cell(r, 5).value or 0, nr.cell(r, 6).value
        c = cfg_apt.get(apt)
        if not c:
            fails.append(f'apartment {apt} is in the workbook but not in the take-off config')
        else:
            near(c['internal'], internal, f'apt {apt} internal')
            near(c['balcony'] or 0, bal, f'apt {apt} balcony')
        has(1, f(internal), f'cover internal {apt}')
        has(1, f(tot), f'cover NSA {apt}')
        p = page_of.get((lv, 'apt'))
        if p:
            has(p, f'{apt}   {f(tot)}', f'{lv} plan label {apt}')
            has(p, f'{apt} {f(tot)}', f'{lv} footer {apt}')
        lvl_tot[lv] = round(lvl_tot.get(lv, 0) + tot, 1)
        r += 1
    has(1, f(nr.cell(r, 6).value), 'cover NSA grand total')
    for lv, t in lvl_tot.items():
        p = page_of.get((lv, 'apt'))
        if p:
            has(p, f'level total {f(t)}', f'{lv} level NSA total')
        has(1, f(t), f'cover NSA {lv}')

    # ---- level NSA block on the workbook agrees with the unit-by-unit sum
    r = 41
    while nr.cell(r, 1).value and str(nr.cell(r, 1).value).upper() != 'TOTAL':
        lv = str(nr.cell(r, 1).value)
        if lv in lvl_tot:
            near(nr.cell(r, 5).value, lvl_tot[lv], f'{lv} NSA-by-level vs unit sum')
        r += 1

    # ---- category sheets reconcile to the Area by Category tab
    acrow = {}
    r = 10
    while ac.cell(r, 1).value and str(ac.cell(r, 1).value).upper() != 'TOTAL':
        acrow[str(ac.cell(r, 1).value)] = r
        r += 1
    for lv, cats in cfg.get('categories', {}).items():
        p = page_of.get((lv, 'cat'))
        for k, v in cats.items():
            if p:
                has(p, f'{v:,.1f}', f'{lv} category {k}')
        if lv in acrow:
            near(round(sum(cats.values()), 1), ac.cell(acrow[lv], 8).value, f'{lv} categories sum vs FECA')

    # ---- GBA authority must appear on the cover and on every measured-area sheet
    for s in ('Property Council of Australia', 'Glossary of Property Terms', 'outside face'):
        has(1, s, 'GBA authority on the cover')
    for (lv, kind), p in page_of.items():
        if kind == 'mk':
            has(p, 'Glossary of Property Terms', f'{lv} GBA authority in the footer')

    # ---- the Summary tab must carry the GBA reference too
    sm = wb['Summary']
    txt = ' '.join(str(c.value) for rw in sm.iter_rows() for c in rw if c.value)
    for s in ('Property Council', 'Glossary of Property Terms',
              'outside face of any enclosing walls', 'Australian Cost Management Manual'):
        checks += 1
        if s not in txt:
            fails.append(f'Summary tab is missing the measurement reference: {s!r}')

    print(f'{checks} checks run')
    if fails:
        print(f'FAILED — {len(fails)} figure(s) do not tie:')
        for x in fails:
            print('  -', x)
        return 1
    print('CLEAN — every figure on the markup ties to the workbook.')
    return 0


if __name__ == '__main__':
    sys.exit(main())
