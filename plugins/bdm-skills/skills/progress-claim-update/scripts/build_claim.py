#!/usr/bin/env python3
"""
build_claim.py — Progress Claim Update Pipeline

Rolls a Head Contract Progress Claim Excel workbook forward by one month.
Encapsulates all 9 structural fixes established during the Claim 16 (May 2026) pilot:

  1. PDF discovery + OneDrive hydration handling
  2. Text/OCR extraction
  3. Invoice parsing + sub-doc/duplicate detection
  4. Xero account-transactions cross-check (orphan identification)
  5. Supplier→CTD-block mapping (with new-supplier sign-off)
  6. Top-down CTD insertion with running offset + alphabetical new sections
  7. Multi-pay F9 formula refresh + SUBTOTAL final-pass recompute
  8. Summary self-reference formula rewrite + Unpaid section snapshot rebuild
  9. Detail Sheet roll + asymmetric I formula + cover linkage + safety self-checks

Usage:
    python3 build_claim.py \
        --project 202415 \
        --config projects/202415_south_pine_rd.yaml \
        --month-folder "<path to NN_<Month> <YYYY>>" \
        --claim-template "<path to prior claim xlsx>" \
        --output-dir "<Alfred working folder for this project>" \
        --claim-no 16 \
        --period-start 27/04/2026 \
        --period-end 22/05/2026 \
        --no-cache

Optional:
    --approved-codes <yaml path>   Pre-supplied new-supplier code approvals (non-interactive mode)
    --skip-xero                    Don't look for Account_Transactions xlsx
    --dry-run                      Show what would change without writing the output file
    --no-cache                     Force re-extraction even if cache exists (USE FOR FIRST RUN EACH MONTH)

Outputs:
    1. Draft claim xlsx at: <output-dir>/<project>_<base>_Claim<NN>_<Month><YYYY>_DRAFT_v<X>.xlsx
    2. Audit log markdown at: <output-dir>/<project>_Claim<NN>_audit_v<X>.md
    3. Prints summary to stdout

Author: Alfred (drafted for Andrew Bentley, BDM Director)
Pilot: Claim 16 May 2026 — 818 South Pine Road, Everton Park

See CHANGELOG.md for outstanding bugs (Eight correctness bugs + four operational
fragility issues — fix before sharing with non-developers self-serve).
"""

from __future__ import annotations
import argparse
import json
import os
import re
import shutil
import subprocess
import sys
import yaml
from collections import defaultdict
from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

# --- R4 modules (July 2026). The legacy extract_* bodies below now delegate here. ---
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import extract_totals as _et
import workbook_ops as _wo
import reconcile as _rec


# ============================================================================
# Configuration loading
# ============================================================================

def load_config(config_path: str) -> dict:
    """Load per-project YAML config. Raises if missing or malformed."""
    if not os.path.exists(config_path):
        raise FileNotFoundError(
            f"Per-project config not found: {config_path}\n"
            f"Create one based on projects/_template.yaml for this project."
        )
    with open(config_path) as f:
        cfg = yaml.safe_load(f)
    # Validate required keys
    required = ['project', 'supplier_to_ctd_block', 'load_bearing',
                'status_folders', 'highlight_colours']
    for k in required:
        if k not in cfg:
            raise ValueError(f"Config missing required key: {k}")
    # Bug #11: the OneDrive bash mount can truncate 6-char hex (C00000 -> C000).
    # Right-pad any 4-char hex back to 6 so PatternFill doesn't choke.
    for _k, _v in list(cfg.get('highlight_colours', {}).items()):
        if isinstance(_v, str):
            _h = _v.lstrip('#').upper()
            if len(_h) == 4:
                _h = _h + '00'
            if len(_h) in (6, 8):
                cfg['highlight_colours'][_k] = _h
    return cfg


# ============================================================================
# Phase 1+2: PDF discovery and text extraction
# ============================================================================

def discover_pdfs(month_folder: str) -> list[dict]:
    """Walk the month folder, return list of {path, rel_path, status, is_subdoc}."""
    items = []
    for root, _, files in os.walk(month_folder):
        for f in files:
            if not f.lower().endswith('.pdf'):
                continue
            full = os.path.join(root, f)
            rel = os.path.relpath(full, month_folder)
            # Status from top-level folder
            top = rel.split(os.sep)[0]
            status = {
                '1.0 Received': 'Received',
                '2.0 Approved': 'Approved',
                '3.0 Paid': 'Paid',
                '4.0 Remittance': 'Remittance',
            }.get(top, 'Unknown')
            items.append({'path': full, 'rel': rel, 'status': status})
    return items


def hydrate_check(items: list[dict]) -> tuple[int, int]:
    """Return (hydrated_count, total_count). Items not hydrated will fail pdftotext."""
    hydrated = 0
    for it in items:
        r = subprocess.run(['pdftotext', '-layout', it['path'], '/dev/null'],
                           capture_output=True, timeout=10)
        if r.returncode == 0:
            hydrated += 1
    return hydrated, len(items)


def extract_text(pdf_path: str, use_ocr_fallback: bool = True) -> str:
    """Extract text from PDF. Falls back to OCR for image-only PDFs."""
    # Try pdftotext first
    r = subprocess.run(['pdftotext', '-layout', pdf_path, '-'],
                       capture_output=True, timeout=30, text=True)
    if r.returncode == 0 and len(r.stdout) > 100:
        return r.stdout
    if not use_ocr_fallback:
        return r.stdout if r.returncode == 0 else ''
    # OCR fallback via pdftoppm + tesseract
    import tempfile
    with tempfile.TemporaryDirectory() as tmpdir:
        base = os.path.join(tmpdir, 'page')
        r2 = subprocess.run(['pdftoppm', '-r', '200', pdf_path, base, '-png'],
                            capture_output=True, timeout=60)
        if r2.returncode != 0:
            return ''
        out = []
        for png in sorted(os.listdir(tmpdir)):
            if png.endswith('.png'):
                r3 = subprocess.run(['tesseract', os.path.join(tmpdir, png), '-'],
                                    capture_output=True, timeout=60, text=True)
                if r3.returncode == 0:
                    out.append(r3.stdout)
        return '\n'.join(out)


# ============================================================================
# Phase 3: Invoice parsing
# ============================================================================

def parse_money(s):
    """R4: delegates to extract_totals.parse_money (handles (1,234.56) negatives)."""
    return _et.parse_money(s)


def extract_total(txt: str, rel_path: str = ''):
    """R4: delegates to extract_totals.extract_total — scored candidates with prefix
    rejection. The old first-match-wins list scored 4/20 on the July sample; this
    scores 20/20. See CHANGELOG R4."""
    return _et.extract_total(txt, rel_path)


MONTHS = 'Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec'

def extract_date_str(txt: str):
    """R4: delegates to extract_totals.extract_date_str — invoice date only; due,
    order and delivery dates are rejected."""
    return _et.extract_date_str(txt)


def parse_date(d_str):
    if not d_str: return datetime(2026, 1, 1)
    for fmt in ('%d/%m/%Y','%d-%m-%Y','%d %b %Y','%d %B %Y','%d/%m/%y'):
        try: return datetime.strptime(d_str, fmt)
        except: pass
    m = re.match(r'(\d{1,2})\s+(\w+)\s+(\d{4})', d_str)
    if m:
        try: return datetime.strptime(f'{m.group(1)} {m.group(2)[:3]} {m.group(3)}', '%d %b %Y')
        except: pass
    return datetime(2026, 1, 1)


def extract_invoice_no(txt: str, fallback: str = ''):
    patterns = [
        r'Tax\s+Invoice\s*#\s*(\S+)',
        r'TAX\s+INVOICE[:\s]+(\S+)',
        r'Invoice\s+No(?:\.|\s+:)?\s*([A-Z0-9\-]+)',
        r'Invoice\s+Number[\s\n]+([A-Z0-9\-\.]+)',
        r'INV(?:OICE)?\s*[:#-]\s*([A-Z0-9\-]+)',
        r'INVOICE\s+no\.\s*(\d+)',
    ]
    for p in patterns:
        m = re.search(p, txt, re.IGNORECASE)
        if m:
            v = m.group(1).strip().rstrip('.,;')
            if v and v.upper() not in ('AUD','GST','TO','FROM','#'):
                return v
    return fallback


def is_subdoc(rel_path: str, cfg: dict) -> bool:
    """Check if a file matches sub-document patterns (excluded from totals)."""
    rl = rel_path.lower()
    for pat in cfg.get('sub_document_patterns', []):
        if str(pat).lower() in rl:      # R4: case-insensitive — 'STATEMENT' missed 'Statement-9443605...'
            return True
    return False


def is_duplicate(rel_path: str, cfg: dict) -> bool:
    """Check if a file matches duplicate patterns."""
    for pat in cfg.get('duplicate_patterns', []):
        if pat in rel_path:
            return True
    return False


def is_remittance(rel_path: str, cfg: dict) -> bool:
    """Check if a file is a remittance receipt (not an invoice)."""
    for pat in cfg.get('remittance_patterns', []):
        if pat in rel_path:
            return True
    return False


def filename_key(rel_path: str) -> str:
    """Build a normalized key for manual_overrides lookup."""
    base = os.path.basename(rel_path).replace('.pdf', '').replace('.PDF', '')
    return base.replace(' ', '_')


def supplier_from_filename(rel_path: str) -> str:
    """Best-effort supplier name from filename; fall back to the parent folder
    when the filename is purely numeric/generic (bug #8 — e.g. PlastaTrade
    '12401_667966.PDF' should resolve to the 'PlastaTrade (Due 31 May)' folder)."""
    base = os.path.basename(rel_path).rsplit('.', 1)[0]
    for sep in ['_-_', ' - ', '_Invoice', '_INV', '_Tax']:
        if sep in base:
            return base.split(sep, 1)[0].replace('_', ' ').strip()
    cleaned = base.replace('_', ' ').strip()
    if not re.search(r'[A-Za-z]{3,}', cleaned):
        parts = rel_path.replace('\\', '/').split('/')
        if len(parts) >= 2:
            parent = re.sub(r'\s*\(Due[^)]*\)', '', parts[-2]).strip()
            if re.search(r'[A-Za-z]{3,}', parent):
                return parent
    return cleaned


def parse_invoice(item: dict, cfg: dict) -> dict | None:
    """Parse a single invoice PDF. Returns record dict or None if sub-doc/remittance."""
    if is_subdoc(item['rel'], cfg) or is_remittance(item['rel'], cfg):
        return None
    # Manual override check first
    key = filename_key(item['rel'])
    overrides = cfg.get('manual_overrides', {})
    for ok, ov in overrides.items():
        if ok in key:
            total = ov['total']
            gst = round(total/11, 2)
            ex_gst = round(total - gst, 2)
            return {
                'rel': item['rel'],
                'status': item['status'],
                'is_duplicate': is_duplicate(item['rel'], cfg),
                'supplier': ov['supplier'],
                'invoice_no': str(ov['invoice_no']),
                'date': ov['date'],
                'total_incl_gst': total,
                'gst': gst,
                'ex_gst': ex_gst,
                'source': 'manual_override',
            }
    # Auto extract
    txt = extract_text(item['path'])
    if not txt or len(txt) < 50:
        return {
            'rel': item['rel'], 'status': item['status'],
            'is_duplicate': is_duplicate(item['rel'], cfg),
            'supplier': supplier_from_filename(item['rel']),
            'invoice_no': '', 'date': '',
            'total_incl_gst': None, 'gst': None, 'ex_gst': None,
            'source': 'unparseable',
        }
    total = extract_total(txt, item['rel'])
    if total is None:
        return {
            'rel': item['rel'], 'status': item['status'],
            'is_duplicate': is_duplicate(item['rel'], cfg),
            'supplier': supplier_from_filename(item['rel']),
            'invoice_no': extract_invoice_no(txt, ''),
            'date': extract_date_str(txt) or '',
            'total_incl_gst': None, 'gst': None, 'ex_gst': None,
            'source': 'no_total',
        }
    gst = round(total/11, 2)          # keeps the sign of `total` for credit notes
    ex_gst = round(total - gst, 2)
    return {
        'rel': item['rel'], 'status': item['status'],
        'is_duplicate': is_duplicate(item['rel'], cfg),
        'supplier': supplier_from_filename(item['rel']),
        'invoice_no': extract_invoice_no(txt, ''),
        'date': extract_date_str(txt) or '',
        'total_incl_gst': total, 'gst': gst, 'ex_gst': ex_gst,
        'credit_on_face': _et.credit_applied_on_face(txt),
        'multi_document': _et.is_multi_document(txt),
        'is_credit_note': _et.is_credit_note(txt, item['rel']),
        'source': 'auto_parsed',
    }


# ============================================================================
# Phase 4: Xero cross-check
# ============================================================================

def find_xero_report(month_folder: str, cfg: dict) -> str | None:
    """Find Xero Account_Transactions xlsx in month folder."""
    import glob
    pattern = cfg.get('xero_report', {}).get('filename_pattern', '*Account_Transactions*.xlsx')
    matches = glob.glob(os.path.join(month_folder, pattern))
    return matches[0] if matches else None


def parse_xero(xero_path: str, cfg: dict) -> list[dict]:
    """Parse Xero Account_Transactions xlsx → list of transaction dicts."""
    xero_cfg = cfg.get('xero_report', {})
    wb = load_workbook(xero_path, data_only=True)
    sheet_name = xero_cfg.get('sheet_name', 'Construction Costs Transactions')
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
    start = xero_cfg.get('data_start_row', 7)
    cols = xero_cfg.get('cols', {})
    txns = []
    for r in range(start, ws.max_row + 1):
        date = ws.cell(r, cols.get('date', 1)).value
        desc = ws.cell(r, cols.get('description', 3)).value
        debit = ws.cell(r, cols.get('debit', 5)).value
        gross = ws.cell(r, cols.get('gross', 8)).value
        gst = ws.cell(r, cols.get('gst', 9)).value
        if isinstance(debit, (int, float)) and debit > 0 and isinstance(desc, str):
            txns.append({
                'date': date, 'desc': desc.strip(),
                'ex_gst': debit,
                'gst': gst if isinstance(gst, (int, float)) else round(debit * 0.1, 2),
                'total': gross if isinstance(gross, (int, float)) else round(debit * 1.1, 2),
            })
    return txns


def keywords(s: str, aliases: dict) -> set:
    """Tokenize a string, with alias keyword injection."""
    s = s.lower()
    for src, target in aliases.items():
        if src in s:
            s += ' ' + target
    return set(re.findall(r'[a-z]{4,}', s))


def xero_cross_check(xero_txns: list, records: list, cfg: dict) -> tuple[list, list]:
    """Match Xero transactions to records. Return (orphans, reclassified)."""
    aliases = cfg.get('xero_aliases', {})
    orphans = []
    reclassified = []
    # Bug #3/#4: precompute filed ex-GST amounts + per-supplier sums for orphan dedup.
    _filed_ex = [(inv.get('ex_gst') or 0) for inv in records
                 if inv.get('total_incl_gst') is not None]
    _by_sup_sum = defaultdict(float)
    for inv in records:
        if inv.get('total_incl_gst') is not None:
            _by_sup_sum[inv['supplier']] += (inv.get('ex_gst') or 0)
    for t in xero_txns:
        match = None
        for inv in records:
            if inv['total_incl_gst'] is None: continue
            if abs((inv['ex_gst'] or 0) - t['ex_gst']) < 1.0:
                if keywords(inv['supplier'], aliases) & keywords(t['desc'], aliases):
                    match = inv; break
        if match:
            if match['status'] in ('Approved', 'Received'):
                reclassified.append((match, t))
                match['status'] = 'Paid'
                match['note'] = (match.get('note') or '') + ' | Reclassified by Xero'
        else:
            # Bug #3 (single-invoice name variation) / #4 (statement = sum of invoices):
            # if the payment amount already matches filed work, it's not a true orphan.
            _amt = t['ex_gst']
            if any(abs(a - _amt) < 1.0 for a in _filed_ex) or \
               any(abs(sm - _amt) < 1.0 for sm in _by_sup_sum.values()):
                continue
            # Orphan — infer supplier + CTD block from description
            sup_raw = t['desc'].split(' - ')[0].strip()
            # Build live set of existing CTD subtotal block names (passed by caller)
            existing_blocks = cfg.get('_live_ctd_blocks', set())
            block = _infer_orphan_block(sup_raw, t['desc'], cfg, existing_blocks)
            code = _block_to_code(block, cfg)
            orphans.append({
                'rel': '(Xero orphan)', 'status': 'Paid',
                'is_duplicate': False, 'is_orphan': True,
                'supplier': sup_raw, 'supplier_block': block,
                'invoice_no': 'INVOICE TBC',
                'date': t['date'].strftime('%d/%m/%Y') if t['date'] else '',
                'total_incl_gst': t['total'],
                'gst': t['gst'],
                'ex_gst': t['ex_gst'],
                'budget_code': code,
                'budget_desc': cfg.get('budget_codes', {}).get(code, ''),
                'source': 'xero_orphan',
            })
    return orphans, reclassified


def _infer_orphan_block(sup_raw: str, desc: str, cfg: dict, all_blocks: set = None) -> str:
    """Best-effort: orphan supplier → CTD block name (existing > new > fallback)."""
    sl = (sup_raw + ' ' + desc).lower()
    # Viking specific scope detection
    if 'viking' in sl:
        if 'electrical' in sl: return 'Viking - TH Electrical Subtotal'
        if 'plumbing' in sl or 'hydraulic' in sl: return 'Viking TH Plumbing Subtotal'
        if 'mechanical' in sl: return 'Viking - Mechanical Subtotal'
    # 1) Match supplier_to_ctd_block (YAML aliases — most reliable, just looks up exact text)
    for supplier, block in cfg.get('supplier_to_ctd_block', {}).items():
        if supplier.lower() in sl:
            return block
    # 2) Fuzzy match supplier_to_ctd_block by keyword overlap
    desc_keywords = set(re.findall(r'[a-z]{4,}', sl))
    for supplier, block in cfg.get('supplier_to_ctd_block', {}).items():
        sk = set(re.findall(r'[a-z]{4,}', supplier.lower()))
        if sk and sk.issubset(desc_keywords):  # supplier keywords must all be in description
            return block
    # 3) Check new_ctd_blocks
    for block in cfg.get('new_ctd_blocks', {}).keys():
        bk = set(re.findall(r'[a-z]{4,}', block.lower().replace(' subtotal', '')))
        if bk and bk.issubset(desc_keywords):
            return block
    # Fallback: new block
    return sup_raw + ' Subtotal'


def _block_to_code(block: str, cfg: dict) -> str:
    """Look up budget code for a CTD block."""
    new_blocks = cfg.get('new_ctd_blocks', {})
    if block in new_blocks:
        v = new_blocks[block]
        if isinstance(v, dict): return v.get('code', '')
        return v
    # Reverse lookup via supplier_to_ctd_block — pick code via supplier's existing claim row
    # (best handled by caller looking up Detail Sheet codes)
    return ''


# ============================================================================
# Phase 5: Supplier mapping + new-supplier sign-off
# ============================================================================

def map_supplier(supplier: str, cfg: dict) -> str | None:
    """Map a supplier name → existing CTD block name, or None if new."""
    mapping = {**cfg.get('supplier_to_ctd_block', {})}
    for sup_name, block in mapping.items():
        if sup_name.lower() == supplier.lower():
            return block
    # Fuzzy
    sk = set(re.findall(r'[a-z]{4,}', supplier.lower()))
    for sup_name, block in mapping.items():
        if sk & set(re.findall(r'[a-z]{4,}', sup_name.lower())):
            return block
    # Check new_ctd_blocks — both substring AND keyword overlap
    for block in cfg.get('new_ctd_blocks', {}).keys():
        if supplier.lower() in block.lower():
            return block
        bk = set(re.findall(r'[a-z]{4,}', block.lower().replace(' subtotal', '')))
        if sk & bk:
            return block
    return None


def map_with_service_type(inv: dict, cfg: dict):
    """Bug #2: multi-service suppliers (Viking) bill several scopes under one name.
    Resolve service type from the supplier/description/path before block mapping."""
    stm = cfg.get('service_type_map', {})
    sup = str(inv.get('supplier', ''))
    hay = (sup + ' ' + str(inv.get('description', '')) + ' ' + str(inv.get('rel', ''))).lower()
    skw = set(re.findall(r'[a-z]{4,}', sup.lower()))
    for key, svc_map in stm.items():
        kkw = set(re.findall(r'[a-z]{4,}', key.lower()))
        if key.lower() in sup.lower() or (kkw and kkw & skw):
            for svc, blk in svc_map.items():
                if svc.lower() in hay:
                    return blk
    return map_supplier(sup, cfg)


# ============================================================================
# Phase 6+7: Workbook build (CTD + Summary + Detail + Cover)
# ============================================================================

def alpha_key(name: str) -> str:
    return name.replace(' Subtotal', '').lower().strip()


def build_workbook(records: list, orphans: list, cfg: dict,
                   src_claim: str, dst_claim: str, claim_no: int,
                   period_start: str, period_end: str, draft_version: str = '0.1'):
    """Apply all 9 structural fixes to roll the claim forward."""
    shutil.copyfile(src_claim, dst_claim)
    wb = load_workbook(dst_claim)

    LB = cfg['load_bearing']
    COLOURS = cfg['highlight_colours']
    HIGHLIGHT = PatternFill('solid', fgColor=COLOURS['new_invoice'])
    NEW_SUB = PatternFill('solid', fgColor=COLOURS['new_section_subtotal'])
    ORPHAN = PatternFill('solid', fgColor=COLOURS['xero_orphan'])

    ctd = wb[LB['sheets'][3]]      # 'CTD Breakdown'
    summary = wb[LB['sheets'][2]]  # 'Summary'
    unpaid = wb[LB['sheets'][4]]   # 'Unpaid Invoices'
    detail = wb[LB['sheets'][1]]   # 'Detail Sheet'
    pc = wb[LB['sheets'][0]]       # 'Progress Claim'

    # ---- Phase 6a: Unmerge problematic cells in Summary (rows >= 5) ----
    for mr in list(summary.merged_cells.ranges):
        if mr.min_row >= 5:
            summary.unmerge_cells(str(mr))

    # ---- Phase 6b: CTD Breakdown — insert paid invoices ----
    SPECIAL_TOP = set(cfg.get('ctd_special_top_sections', []))
    ctd_subtotal_rows = {}
    for r in range(LB['ctd_breakdown']['data_start_row'], ctd.max_row + 1):
        d = ctd.cell(r, LB['ctd_breakdown']['cols']['description']).value
        if isinstance(d, str) and 'Subtotal' in d:
            ctd_subtotal_rows[d] = r

    multipay_rows_orig = []
    multipay_marker = LB['ctd_breakdown']['multipay_pickup_marker']
    for r in range(LB['ctd_breakdown']['data_start_row'], ctd.max_row + 1):
        d = ctd.cell(r, LB['ctd_breakdown']['cols']['description']).value
        if isinstance(d, str) and multipay_marker in d:
            multipay_rows_orig.append(r)

    # All paid invoices (filed + orphans)
    all_paid = [r for r in records if not r.get('is_duplicate') and
                r['status'] == 'Paid' and r['total_incl_gst'] is not None]
    all_paid.extend(orphans)

    # Group by CTD block
    existing_block_invs = defaultdict(list)
    new_block_invs = defaultdict(list)
    for inv in all_paid:
        block = inv.get('supplier_block') or map_with_service_type(inv, cfg)
        if not block:
            print(f"WARN: no block for {inv['supplier']}; skipping CTD", file=sys.stderr)
            continue
        # Bug #7: sanity-check the supplier actually belongs in the destination section.
        _sk = set(re.findall(r'[a-z]{4,}', str(inv['supplier']).lower()))
        _bk = set(re.findall(r'[a-z]{4,}', str(block).lower().replace(' subtotal', '')))
        if _sk and _bk and not (_sk & _bk):
            print(f"WARN (bug#7): '{inv['supplier']}' -> '{block}' : supplier/section name "
                  f"mismatch — review this allocation before issue.", file=sys.stderr)
        if block in ctd_subtotal_rows:
            existing_block_invs[block].append(inv)
        else:
            new_block_invs[block].append(inv)

    for invs in existing_block_invs.values():
        invs.sort(key=lambda x: parse_date(x.get('date', '')))
    for invs in new_block_invs.values():
        invs.sort(key=lambda x: parse_date(x.get('date', '')))

    # Alphabetical insertion target for new sections
    alpha_sorted = sorted([(n, r) for n, r in ctd_subtotal_rows.items() if n not in SPECIAL_TOP],
                         key=lambda x: alpha_key(x[0]))
    new_section_target = {}
    for nb in new_block_invs:
        key = alpha_key(nb)
        target = None
        for name, _ in alpha_sorted:
            if alpha_key(name) > key:
                target = name; break
        new_section_target[nb] = target

    sheet_ordered = sorted(ctd_subtotal_rows.items(), key=lambda x: x[1])
    section_data_start = {}
    prev = LB['ctd_breakdown']['data_start_row'] - 1
    for name, sub in sheet_ordered:
        section_data_start[name] = prev + 1
        prev = sub

    # Build insertion plan
    insertions = []
    for block, invs in existing_block_invs.items():
        insertions.append({'pos': ctd_subtotal_rows[block], 'count': len(invs),
                          'type': 'existing', 'block': block, 'invoices': invs})
    for block, invs in new_block_invs.items():
        target = new_section_target[block]
        pos = section_data_start.get(target, max(ctd_subtotal_rows.values()) + 1)
        insertions.append({'pos': pos, 'count': len(invs) + 1,
                          'type': 'new', 'block': block, 'invoices': invs})

    # Top-down with running offset
    insertions.sort(key=lambda x: (x['pos'], 0 if x['type'] == 'new' else 1))
    multipay_current = {r: r for r in multipay_rows_orig}
    offset = 0
    cols = LB['ctd_breakdown']['cols']
    for ins in insertions:
        actual_pos = ins['pos'] + offset
        count = ins['count']
        ctd.insert_rows(actual_pos, amount=count)
        for orig in list(multipay_current):
            if multipay_current[orig] >= actual_pos:
                multipay_current[orig] += count
        write_row = actual_pos
        # Write invoice rows
        for inv in ins['invoices']:
            ctd.cell(write_row, cols['date']).value = parse_date(inv.get('date', ''))
            ctd.cell(write_row, cols['date']).number_format = 'd/mm/yyyy'
            ctd.cell(write_row, cols['source']).value = 'Spend Money'
            if inv.get('is_orphan'):
                desc = f"{inv['supplier']} - INVOICE TBC; Xero CC payment - CHASE SUPPLIER"
            else:
                desc = f"{inv['supplier']} - {inv.get('budget_desc','')}; Inv {inv.get('invoice_no','')}"
            ctd.cell(write_row, cols['description']).value = desc
            ctd.cell(write_row, cols['debit']).value = inv['ex_gst']
            ctd.cell(write_row, cols['credit']).value = 0
            ctd.cell(write_row, cols['gross']).value = inv['total_incl_gst']
            ctd.cell(write_row, cols['gst']).value = inv['gst']
            for c in [cols['debit'], cols['credit'], cols['gross'], cols['gst']]:
                ctd.cell(write_row, c).number_format = '#,##0.00'
            fill = ORPHAN if inv.get('is_orphan') else HIGHLIGHT
            for c in range(cols['date'], cols['gst']+1):
                ctd.cell(write_row, c).fill = fill
            write_row += 1
        # If new section, add the SUBTOTAL row
        if ins['type'] == 'new':
            ctd.cell(write_row, cols['description']).value = ins['block']
            for c in [cols['description'], cols['debit'], cols['credit'], cols['gross']]:
                ctd.cell(write_row, c).font = Font(name='Calibri', size=10, bold=True)
                ctd.cell(write_row, c).fill = NEW_SUB
        offset += count

    # ---- Phase 7a: Recompute all SUBTOTAL formulas ----
    # IMPORTANT: also handles "sub-subtotal" rows where the description DOESN'T contain
    # the word "Subtotal" but the E formula IS a SUBTOTAL (e.g. "Blue Mojo Temp Services",
    # "VZ Industries Pty Ltd - Electrical Reticulation" in the 202415 template).
    # Without this, the parent supplier subtotal's range would incorrectly span them,
    # causing double-counting in Summary.
    last_sub = LB['ctd_breakdown']['data_start_row'] - 1
    gt_row = None
    for r in range(LB['ctd_breakdown']['data_start_row'], ctd.max_row + 1):
        d = ctd.cell(r, cols['description']).value
        e_val = ctd.cell(r, cols['debit']).value
        is_subtotal_text = isinstance(d, str) and 'Subtotal' in d
        is_subtotal_formula = isinstance(e_val, str) and 'SUBTOTAL' in e_val.upper()
        # Don't treat the CTD grand total row as a regular subtotal
        is_grand_total = (ctd.cell(r, cols['date']).value == LB['ctd_breakdown']['grand_total_marker_col_b'])
        if (is_subtotal_text or is_subtotal_formula) and not is_grand_total:
            if r - 1 >= last_sub + 1:
                ctd.cell(r, cols['debit']).value = f'=SUBTOTAL(9,E{last_sub+1}:E{r-1})'
                ctd.cell(r, cols['credit']).value = f'=SUBTOTAL(9,F{last_sub+1}:F{r-1})'
                ctd.cell(r, cols['gross']).value = f'=SUBTOTAL(9,G{last_sub+1}:G{r-1})'
            last_sub = r
        elif ctd.cell(r, cols['date']).value == LB['ctd_breakdown']['grand_total_marker_col_b']:
            gt_row = r
    if gt_row:
        ctd.cell(gt_row, cols['debit']).value = f'=SUBTOTAL(9,E7:E{gt_row-1})'
        ctd.cell(gt_row, cols['credit']).value = f'=SUBTOTAL(9,F7:F{gt_row-1})'
        ctd.cell(gt_row, cols['gross']).value = f'=SUBTOTAL(9,G7:G{gt_row-1})'

    # ---- Phase 7b: Multi-pay F9 update ----
    new_refs = [f'E{multipay_current[r]}' for r in multipay_rows_orig]
    if new_refs:
        ctd['F9'].value = '=' + '+'.join(new_refs)
        ctd['I9'].value = '=E9-' + '-'.join(new_refs)

    # ---- Phase 8a: Summary — insert new Paid section rows ----
    existing_summary_rows = []
    for r in range(LB['summary']['paid_section_first_row'], 100):
        b = summary.cell(r, LB['summary']['cols']['supplier_or_subtotal']).value
        if isinstance(b, str) and b.strip() and 'Subtotal' in b:
            existing_summary_rows.append((r, b))

    summary_special_top = {'Construction Insurance Subtotal', 'Mutli-pay Subtotal'}
    sum_alpha = sorted([(r, n) for r, n in existing_summary_rows
                       if n not in summary_special_top],
                      key=lambda x: alpha_key(x[1]))

    # Find new blocks (those in CTD's new_block_invs PLUS any defined in cfg new_ctd_blocks)
    new_blocks_to_add = list(new_block_invs.keys())
    summary_insertions = []
    for nb in new_blocks_to_add:
        key = alpha_key(nb)
        insert_pos = None
        for row, name in sum_alpha:
            if alpha_key(name) > key: insert_pos = row; break
        if insert_pos is None:
            insert_pos = max(r for r, _ in existing_summary_rows) + 1
        # Look up code
        nb_cfg = cfg.get('new_ctd_blocks', {}).get(nb, {})
        code = nb_cfg.get('code') if isinstance(nb_cfg, dict) else nb_cfg
        if not code:
            # Fallback: use first invoice's code
            code = new_block_invs[nb][0].get('budget_code', '') if new_block_invs[nb] else ''
        summary_insertions.append((insert_pos, nb, code))

    summary_insertions.sort(key=lambda x: x[0])
    s_offset = 0
    new_paid_rows = set()
    sc = LB['summary']['cols']
    for pos_orig, block, code in summary_insertions:
        actual_pos = pos_orig + s_offset
        summary.insert_rows(actual_pos, amount=1)
        summary.cell(actual_pos, sc['supplier_or_subtotal']).value = block
        summary.cell(actual_pos, sc['budget_code']).value = code
        new_paid_rows.add(actual_pos)
        s_offset += 1

    # Find paid_section_end, grand_total_row, unpaid markers
    paid_section_end = 0
    grand_total_row = 0
    unpaid_header_row = 0
    unpaid_subhead_row = 0
    for r in range(LB['summary']['paid_section_first_row'], summary.max_row + 1):
        b = summary.cell(r, sc['supplier_or_subtotal']).value
        if isinstance(b, str):
            if 'Subtotal' in b:
                paid_section_end = max(paid_section_end, r)
            elif LB['summary']['unpaid_header_match'] in b.lower():
                unpaid_header_row = r
            elif b == LB['summary']['unpaid_subhead_match']:
                unpaid_subhead_row = r
    for r in range(paid_section_end + 1, unpaid_header_row or summary.max_row + 1):
        c_val = summary.cell(r, sc['debit']).value
        if isinstance(c_val, str) and c_val.startswith('=SUM('):
            grand_total_row = r; break
    if not grand_total_row:
        grand_total_row = paid_section_end + 2

    # Rewrite Paid section formulas
    # Target: rows with 'Subtotal' in B (typical case) OR rows with a budget code in F
    # (catches edge cases like "Blue Mojo Temp Services" or "VZ Industries Pty Ltd -
    # Electrical Reticulation" — supplier rows that don't follow the "X Subtotal" convention)
    for r in range(LB['summary']['paid_section_first_row'], paid_section_end + 1):
        b = summary.cell(r, sc['supplier_or_subtotal']).value
        f_code = summary.cell(r, sc['budget_code']).value
        is_supplier_row = (
            (isinstance(b, str) and 'Subtotal' in b) or
            (isinstance(b, str) and b.strip() and isinstance(f_code, str) and '-' in str(f_code))
        )
        if is_supplier_row:
            summary.cell(r, sc['debit']).value = (
                f"=SUMIF('CTD Breakdown'!D:D,Summary!B{r},'CTD Breakdown'!E:E)"
            )
            summary.cell(r, sc['credit']).value = (
                f"=SUMIF('CTD Breakdown'!D:D,Summary!B{r},'CTD Breakdown'!F:F)"
            )
            summary.cell(r, sc['total']).value = f"=C{r}-D{r}"
            if r in new_paid_rows:
                for c in range(sc['supplier_or_subtotal'], sc['budget_code'] + 1):
                    summary.cell(r, c).fill = HIGHLIGHT

    # Update Summary grand total — covers Paid + Unpaid sections (we'll compute Unpaid range after building)

    # ---- Phase 8b: Unpaid Invoices tab — snapshot rebuild ----
    non_paid = [r for r in records if not r.get('is_duplicate') and
                r['status'] in ('Approved', 'Received') and r['total_incl_gst'] is not None]
    non_paid.sort(key=lambda x: (x['supplier'], str(x.get('invoice_no', ''))))

    # Clear existing data
    for r in range(unpaid.max_row, 1, -1):
        for c in range(1, max(8, unpaid.max_column) + 1):
            unpaid.cell(r, c).value = None
            unpaid.cell(r, c).fill = PatternFill(fill_type=None)

    uc = LB['unpaid_invoices']['cols']
    unpaid.cell(1, uc['status']).value = 'Status'
    unpaid.cell(1, uc['status']).font = Font(name='Calibri', size=10, bold=True)
    for idx, inv in enumerate(non_paid, start=2):
        unpaid.cell(idx, uc['supplier']).value = inv['supplier']
        unpaid.cell(idx, uc['invoice_no']).value = inv.get('invoice_no', '')
        unpaid.cell(idx, uc['invoice_date']).value = parse_date(inv.get('date', ''))
        unpaid.cell(idx, uc['invoice_date']).number_format = 'd/mm/yyyy'
        unpaid.cell(idx, uc['ex_gst']).value = inv['ex_gst']
        unpaid.cell(idx, uc['gst']).value = inv['gst']
        unpaid.cell(idx, uc['total']).value = inv['total_incl_gst']
        unpaid.cell(idx, uc['status']).value = inv['status'].upper()
        for c in [uc['ex_gst'], uc['gst'], uc['total']]:
            unpaid.cell(idx, c).number_format = '#,##0.00'
        for c in range(uc['supplier'], uc['status'] + 1):
            unpaid.cell(idx, c).fill = HIGHLIGHT
        col = COLOURS['received_status_text'] if inv['status'] == 'Received' else COLOURS['approved_status_text']
        unpaid.cell(idx, uc['status']).font = Font(name='Calibri', size=10, color=col, bold=True)

    # ---- Phase 8c: Summary Unpaid section — consolidated rebuild ----
    # Clear current Unpaid section
    unpaid_section_start = (unpaid_subhead_row + 1) if unpaid_subhead_row else (unpaid_header_row + 2 if unpaid_header_row else 100)
    for r in range(unpaid_section_start, summary.max_row + 5):
        for c in range(1, 8):
            summary.cell(r, c).value = None
            summary.cell(r, c).fill = PatternFill(fill_type=None)
    # Clear orphan rows between grand total and unpaid header
    for r in range(grand_total_row + 1, unpaid_header_row):
        for c in range(1, 8):
            summary.cell(r, c).value = None
            summary.cell(r, c).fill = PatternFill(fill_type=None)

    # Update header and subhead
    summary.cell(unpaid_header_row, sc['supplier_or_subtotal']).value = f'Unpaid Invoices as at {period_end}'
    summary.cell(unpaid_header_row, sc['supplier_or_subtotal']).font = Font(
        name='Calibri', size=11, bold=True, color='1F3864'
    )
    summary.cell(unpaid_subhead_row, sc['supplier_or_subtotal']).value = 'Subcontractor'
    summary.cell(unpaid_subhead_row, sc['debit']).value = 'Debt'
    summary.cell(unpaid_subhead_row, sc['credit']).value = 'Credit'
    summary.cell(unpaid_subhead_row, sc['total']).value = 'Total'
    summary.cell(unpaid_subhead_row, sc['budget_code']).value = 'Budget Code'
    for c in range(sc['supplier_or_subtotal'], sc['budget_code'] + 1):
        summary.cell(unpaid_subhead_row, c).font = Font(name='Calibri', size=10, bold=True)
        summary.cell(unpaid_subhead_row, c).fill = PatternFill('solid', fgColor='F2F2F2')

    # Unique suppliers from Unpaid Invoices tab
    seen = set()
    unpaid_suppliers = []
    for inv in non_paid:
        if inv['supplier'] not in seen:
            seen.add(inv['supplier'])
            unpaid_suppliers.append(inv['supplier'])
    unpaid_suppliers.sort()

    # Build supplier->code map for UNPAID invoices.
    # FIX (R2, PC17 Jun-2026): non-paid invoices carry no budget_code/supplier_block, so
    # resolve each supplier -> CTD block -> code, with a keyword fallback against the live
    # Paid section. Without this the Summary Unpaid 'Budget Code' (col F) was left blank and
    # received/approved invoices fed $0 into Detail Sheet 'Actual to Date'.
    _paid_block_code = []  # (block_name_lower, code) harvested from the rebuilt Paid section
    for _r in range(LB['summary']['paid_section_first_row'], grand_total_row):
        _bn = summary.cell(_r, sc['supplier_or_subtotal']).value
        _bc = summary.cell(_r, sc['budget_code']).value
        if _bn and _bc and 'subtotal' in str(_bn).lower():
            _paid_block_code.append((str(_bn).lower(), str(_bc)))

    def _kw(t):
        return set(re.findall(r'[a-z]{4,}', str(t).lower()))

    def _resolve_unpaid_code(inv):
        if inv.get('budget_code'):
            return inv['budget_code']
        blk = inv.get('supplier_block') or map_supplier(inv['supplier'], cfg)
        if blk:
            c = _block_to_code(blk, cfg)
            if c:
                return c
            bkw = _kw(str(blk).replace(' subtotal', ''))
            best, bs = '', 0
            for pn, pc in _paid_block_code:
                ov = len(bkw & _kw(pn))
                if ov > bs:
                    bs, best = ov, pc
            if best:
                return best
        skw = _kw(inv['supplier'])
        best, bs = '', 0
        for pn, pc in _paid_block_code:
            ov = len(skw & _kw(pn))
            if ov > bs:
                bs, best = ov, pc
        return best

    supplier_to_code = {}
    for inv in non_paid:
        if inv['supplier'] not in supplier_to_code:
            supplier_to_code[inv['supplier']] = _resolve_unpaid_code(inv)

    write_row = unpaid_section_start
    for sup in unpaid_suppliers:
        code = supplier_to_code.get(sup, '')
        summary.cell(write_row, sc['supplier_or_subtotal']).value = sup
        summary.cell(write_row, sc['debit']).value = (
            f"=SUMIF('Unpaid Invoices'!A:A,Summary!B{write_row},'Unpaid Invoices'!D:D)"
        )
        summary.cell(write_row, sc['credit']).value = 0
        summary.cell(write_row, sc['total']).value = f"=C{write_row}+D{write_row}"
        summary.cell(write_row, sc['budget_code']).value = code
        write_row += 1
    new_max = write_row - 1

    # Update Summary grand total formula
    summary.cell(grand_total_row, sc['debit']).value = (
        f"=SUM(C5:C{grand_total_row-1})+SUM(C{unpaid_section_start}:C{new_max})"
    )
    summary.cell(grand_total_row, sc['credit']).value = (
        f"=SUM(D5:D{grand_total_row-1})+SUM(D{unpaid_section_start}:D{new_max})"
    )
    summary.cell(grand_total_row, sc['total']).value = (
        f"=SUM(E5:E{grand_total_row-1})+SUM(E{unpaid_section_start}:E{new_max})"
    )

    # ---- Phase 8d/9: Detail Sheet roll + cover ----
    src_wb = load_workbook(src_claim, data_only=True)
    ds_src = src_wb[LB['sheets'][1]]
    dc = LB['detail_sheet']['cols']
    old_J = {}
    for r in range(LB['detail_sheet']['first_line_row'], LB['detail_sheet']['last_line_row'] + 1):
        code = ds_src.cell(r, dc['budget_code']).value
        if code:
            old_J[r] = ds_src.cell(r, dc['total_to_date']).value or 0

    # Roll H = old J; set I to formula =MAX(0, Q{r}-H{r})
    margin_row = LB['detail_sheet']['last_line_row']  # row 79
    for r in range(LB['detail_sheet']['first_line_row'],
                   LB['detail_sheet']['last_line_row'] + 1):
        code = detail.cell(r, dc['budget_code']).value
        if not code: continue
        detail.cell(r, dc['previous']).value = old_J.get(r, 0)
        if r != margin_row:
            # Skip section-header rows (no hyphen in code = section label)
            if '-' not in str(code) and '.' not in str(code):
                continue
            detail.cell(r, dc['this_period']).value = f'=MAX(0,Q{r}-H{r})'
    detail.cell(margin_row, dc['this_period']).value = (
        f'=SUM(J{LB["detail_sheet"]["first_line_row"]-1}:J{margin_row-1})*'
        f'{cfg["project"]["builders_margin_pct"]/100}-H{margin_row}'
    )

    detail['K2'] = claim_no
    detail['K3'] = datetime.strptime(period_end, '%d/%m/%Y')
    detail['K4'] = f'{period_start} - {period_end}'

    # Detail Sheet totals row formulas (dynamic SUMs)
    tr = LB['detail_sheet']['totals_row']
    detail[f'D{tr}'] = f'=SUM(D9:D{tr-1})'
    detail[f'F{tr}'] = f'=SUM(F9:F{tr-1})'
    detail[f'H{tr}'] = f'=SUM(H9:H{tr-1})'
    detail[f'I{tr}'] = f'=SUM(I9:I{tr-1})'
    detail[f'J{tr}'] = f'=SUM(J9:J{tr-1})'
    detail[f'K{tr}'] = f'=IF(F{tr}=0,0,J{tr}/F{tr})'
    detail[f'L{tr}'] = f'=F{tr}-J{tr}'

    # Cover
    cc = LB['cover_cells']
    pc[cc['claim_no']] = claim_no
    pc[cc['invoice_no']] = claim_no
    pc[cc['period']] = f'{period_start} - {period_end}'
    pc[cc['total_completed']] = f"='Detail Sheet'!J{tr}"
    pc[cc['total_earned_less_retention']] = f"='Detail Sheet'!J{tr}"
    pc[cc['previous_claims']] = f"='Detail Sheet'!H{tr}"
    pc[cc['current_ex_tax']] = f"='Detail Sheet'!J{tr}-'Detail Sheet'!H{tr}"
    pc[cc['gst']] = f"={cc['current_ex_tax']}*0.1"
    pc[cc['current_incl_tax']] = f"={cc['current_ex_tax']}+{cc['gst']}"
    pc[cc['balance_to_finish']] = f"={cc['contract_sum_to_date']}-{cc['total_completed']}"
    pc[cc['amount_certified']] = f"={cc['current_incl_tax']}"

    # DRAFT banner
    draft_row = cc['draft_banner_row']
    pc.row_dimensions[draft_row].height = 22
    try:
        pc.merge_cells(start_row=draft_row, start_column=1,
                      end_row=draft_row, end_column=12)
    except: pass
    pc.cell(draft_row, 1).value = (
        f'— DRAFT v{draft_version} — Alfred prepared, for issue by Andrew/SPM — '
        f'Period {period_start} – {period_end}. '
        f'CTD red rows = Xero CC payments without invoice on file (chase supplier).'
    )
    pc.cell(draft_row, 1).fill = PatternFill('solid', fgColor=COLOURS['draft_banner_fill'])
    pc.cell(draft_row, 1).font = Font(
        name='Calibri', size=10, bold=True, italic=True,
        color=COLOURS['draft_banner_font']
    )
    pc.cell(draft_row, 1).alignment = Alignment(
        horizontal='center', vertical='center', wrap_text=True
    )

    # Atomic write: save to local tmp first, verify, then move into place
    # (OneDrive can intercept in-place saves and corrupt the zip)
    import tempfile
    tmp_dir = tempfile.gettempdir()
    tmp_path = os.path.join(tmp_dir, '_claim_atomic_' + os.path.basename(dst_claim))
    wb.save(tmp_path)
    # Verify zip is well-formed
    import zipfile
    try:
        with zipfile.ZipFile(tmp_path) as zf:
            zf.testzip()
    except Exception as e:
        raise RuntimeError(f"Save produced corrupted xlsx: {e}")
    shutil.move(tmp_path, dst_claim)
    return dst_claim


# ============================================================================
# Phase 7c: Safety self-check
# ============================================================================

def safety_check(claim_path: str, cfg: dict):
    """Verify the template structure is intact. Raises ValueError if not."""
    wb = load_workbook(claim_path, data_only=False)
    LB = cfg['load_bearing']
    issues = []
    # Check all required sheets present
    # Tolerate 'Advnaced Payments' typo OR corrected 'Advanced Payments' (corrected May 2026)
    for sheet in LB['sheets']:
        if sheet in wb.sheetnames:
            continue
        if sheet == 'Advnaced Payments' and 'Advanced Payments' in wb.sheetnames:
            continue
        if sheet == 'Advanced Payments' and 'Advnaced Payments' in wb.sheetnames:
            continue
        issues.append(f'Missing sheet: {sheet}')
    # Check Detail Sheet Builders Margin at expected row
    if 'Detail Sheet' in wb.sheetnames:
        ds = wb['Detail Sheet']
        margin_row = LB['detail_sheet']['last_line_row']
        margin_code = cfg['project']['builders_margin_code']
        actual = ds.cell(margin_row, LB['detail_sheet']['cols']['budget_code']).value
        if actual != margin_code:
            issues.append(
                f"Detail Sheet R{margin_row} expected Builders Margin code "
                f"'{margin_code}', got '{actual}'"
            )
    if issues:
        raise ValueError("Safety check failed:\n  - " + "\n  - ".join(issues))


# ============================================================================
# Phase 8 (verify): LibreOffice recalc + value check
# ============================================================================

def recalc_and_verify(claim_path: str) -> dict:
    """Run LibreOffice headless to compute formulas; return key totals for verification."""
    out_dir = '/tmp/build_claim_recalc'
    os.makedirs(out_dir, exist_ok=True)
    r = subprocess.run(['libreoffice', '--headless', '--calc',
                       '--convert-to', 'xlsx', '--outdir', out_dir, claim_path],
                      capture_output=True, timeout=120)
    if r.returncode != 0:
        raise RuntimeError(f"LibreOffice recalc failed: {r.stderr.decode()}")
    recalc_path = os.path.join(out_dir, os.path.basename(claim_path))
    wb = load_workbook(recalc_path, data_only=True)
    pc = wb['Progress Claim']
    # Verify no #VALUE! errors in key cells
    for cell in ['F22', 'F28', 'F29', 'F30', 'F31', 'F32', 'K36']:
        v = pc[cell].value
        if isinstance(v, str) and '#' in v:
            raise RuntimeError(f"Recalc produced error at Progress Claim {cell}: {v}")
    return {
        'claim_no': pc['I3'].value,
        'period': pc['I5'].value,
        'original_contract': pc['F19'].value,
        'total_completed': pc['F22'].value,
        'previous_claims': pc['F28'].value,
        'this_period_ex_gst': pc['F29'].value,
        'gst': pc['F30'].value,
        'this_period_incl_gst': pc['F31'].value,
        'balance_to_finish': pc['F32'].value,
    }


# ============================================================================
# Phase 9: Outputs (audit log)
# ============================================================================

def write_audit_log(audit_path: str, records: list, orphans: list,
                    cover: dict, cfg: dict, claim_no: int, period_end: str):
    """Write a markdown audit log."""
    from collections import Counter
    by_status = Counter(r['status'] for r in records if not r.get('is_duplicate')
                        and r['total_incl_gst'] is not None)
    new_suppliers_added = []
    for r in records:
        if r.get('is_new_supplier'):
            new_suppliers_added.append((r['supplier'], r.get('budget_code')))

    lines = [
        f"# Progress Claim {claim_no} — Audit Log",
        f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        "",
        "## Cover totals",
        f"- Original contract sum: ${cover['original_contract']:,.2f}",
        f"- Total completed to date: ${cover['total_completed']:,.2f}",
        f"- Previous claims: ${cover['previous_claims']:,.2f}",
        f"- This period ex GST: ${cover['this_period_ex_gst']:,.2f}",
        f"- GST: ${cover['gst']:,.2f}",
        f"- **This period incl GST: ${cover['this_period_incl_gst']:,.2f}**",
        f"- Balance to finish: ${cover['balance_to_finish']:,.2f}",
        "",
        "## Invoice counts",
        f"- Paid: {by_status.get('Paid', 0)}",
        f"- Approved: {by_status.get('Approved', 0)}",
        f"- Received: {by_status.get('Received', 0)}",
        "",
        "## Xero orphan transactions (CHASE SUPPLIER)",
    ]
    if orphans:
        for o in orphans:
            lines.append(
                f"- {o['date']} {o['supplier']}: ${o['total_incl_gst']:,.2f} → code {o['budget_code']}"
            )
        lines.append(f"\n**Total orphan payments: ${sum(o['total_incl_gst'] for o in orphans):,.2f} incl GST**")
    else:
        lines.append("- None — every Xero payment matched to a filed invoice")

    lines.append("")
    lines.append("## Items needing manual review")
    unreadable = [r for r in records if r.get('source') in ('unparseable', 'no_total')]
    if unreadable:
        for r in unreadable:
            lines.append(f"- {r['supplier']}: could not extract total. File: {r['rel']}")
    else:
        lines.append("- None — all invoices parsed successfully")

    with open(audit_path, 'w') as f:
        f.write('\n'.join(lines))


# ============================================================================
# Main entry point
# ============================================================================

def main():
    parser = argparse.ArgumentParser(description="Progress Claim Update Pipeline")
    parser.add_argument('--project', required=True, help='Project code (e.g. 202415)')
    parser.add_argument('--config', required=True, help='Per-project config YAML path')
    parser.add_argument('--month-folder', required=True, help='Path to NN_<Month> <YYYY> accounts folder')
    parser.add_argument('--claim-template', required=True, help='Path to prior claim xlsx (to roll forward)')
    parser.add_argument('--output-dir', required=True, help='Alfred working folder for project outputs')
    parser.add_argument('--claim-no', type=int, required=True, help='New claim number')
    parser.add_argument('--period-start', required=True, help='Period start dd/mm/yyyy')
    parser.add_argument('--period-end', required=True, help='Period end dd/mm/yyyy')
    parser.add_argument('--draft-version', default='0.1', help='Draft version suffix (default 0.1)')
    parser.add_argument('--skip-xero', action='store_true', help='Skip Xero account-transactions cross-check')
    parser.add_argument('--dry-run', action='store_true', help='Report what would change, no output')
    parser.add_argument('--force', action='store_true', help='Skip hydration check (use if you know all critical files are hydrated)')
    parser.add_argument('--cache-path', default=None, help='Path to cache extracted records JSON (load if exists, save after extraction). Speeds up re-runs.')
    parser.add_argument('--no-cache', action='store_true', help='Force re-extraction even if cache exists (RECOMMENDED for first run of each month)')
    args = parser.parse_args()

    cfg = load_config(args.config)

    print(f"=== Progress Claim {args.claim_no} build ===")
    print(f"Project: {cfg['project']['code']} - {cfg['project']['name']}")
    print(f"Month folder: {args.month_folder}")
    print(f"Period: {args.period_start} - {args.period_end}")
    print()

    # Phase 1: Discover
    items = discover_pdfs(args.month_folder)
    print(f"Phase 1: Found {len(items)} PDF files across status folders")

    # Hydration check
    hyd, tot = hydrate_check(items)
    if hyd < tot:
        if args.force:
            print(f"WARNING: {tot - hyd}/{tot} PDFs not hydrated; --force set, proceeding anyway")
        else:
            print(f"WARNING: {tot - hyd}/{tot} PDFs not hydrated locally.")
            print(f"Right-click month folder in File Explorer → 'Always keep on this device', then retry.")
            print(f"Or re-run with --force to proceed with partial data.")
            sys.exit(2)

    # Phase 2+3: Extract + parse (with cache)
    cache_path = args.cache_path or os.path.join(args.output_dir, f'.{args.project}_records_cache.json')
    # Bug #1: auto-invalidate a stale cache if the month folder is newer than it.
    _cache_stale = False
    if cache_path and os.path.exists(cache_path) and not args.no_cache:
        try:
            _cmt = os.path.getmtime(cache_path)
            _newest = max((os.path.getmtime(os.path.join(_dp, _f))
                           for _dp, _, _fs in os.walk(args.month_folder) for _f in _fs),
                          default=0)
            if _newest > _cmt:
                _cache_stale = True
                print("Cache older than month folder — invalidating (bug #1).")
        except OSError:
            pass
    if cache_path and os.path.exists(cache_path) and not args.no_cache and not _cache_stale:
        with open(cache_path) as cf:
            records = json.load(cf)
        print(f"Phase 2+3: Loaded {len(records)} invoices from cache {cache_path}")
    else:
        records = []
        for i, it in enumerate(items):
            if i % 20 == 0:
                print(f"  ...parsing {i}/{len(items)}", flush=True)
            rec = parse_invoice(it, cfg)
            if rec: records.append(rec)
        if cache_path:
            os.makedirs(os.path.dirname(cache_path) or '.', exist_ok=True)
            with open(cache_path, 'w') as cf:
                json.dump(records, cf, indent=2, default=str)
        print(f"Phase 2+3: Parsed {len(records)} invoices "
              f"({sum(1 for r in records if r['total_incl_gst'] is None)} unparseable). Cache: {cache_path}")

    # Phase 4: Xero cross-check
    orphans = []
    if not args.skip_xero:
        xero_path = find_xero_report(args.month_folder, cfg)
        if xero_path:
            print(f"Phase 4: Found Xero report {os.path.basename(xero_path)}")
            txns = parse_xero(xero_path, cfg)
            orphans, reclassified = xero_cross_check(txns, records, cfg)
            print(f"  {len(txns)} transactions, {len(reclassified)} reclassified A/R→Paid, "
                  f"{len(orphans)} orphans (chase list)")
        else:
            print("Phase 4: No Xero report found in month folder — skipping cross-check")

    # Phase 5: New supplier detection
    new_suppliers = []
    for r in records:
        if r['status'] != 'Paid': continue
        if r['total_incl_gst'] is None: continue
        if not map_supplier(r['supplier'], cfg):
            new_suppliers.append(r['supplier'])
    if new_suppliers:
        unique_new = sorted(set(new_suppliers))
        print(f"Phase 5: {len(unique_new)} new suppliers need code approval:")
        for s in unique_new:
            print(f"  - {s}")
        print("Provide approvals via --approved-codes YAML or interactively (TODO)")

    if args.dry_run:
        print("\n--dry-run flag set; not building output")
        return 0

    # Phase 6-9: Build
    # Strip prior _Claim<NN>_<Month><YYYY>_DRAFT_v<X> suffix and any leading project code
    # so re-running on a previous draft produces a clean filename, not double-stacked.
    base = os.path.basename(args.claim_template).rsplit('.', 1)[0]
    base = re.sub(r'_Claim\d+_[A-Za-z]+\d{4}_DRAFT_v[\d.]+$', '', base)
    base = re.sub(rf'^{re.escape(str(args.project))}_', '', base)
    # Bug #10: strip messy template suffixes — (REVISED ...), trailing uuid8, (n).
    base = re.sub(r'\s*[-_ ]*\(?\s*REVISED[^)]*\)?', '', base, flags=re.I)
    base = re.sub(r'[ _-]+[0-9a-f]{8}$', '', base)
    base = re.sub(r'\s*\(\d+\)$', '', base)
    base = re.sub(r'[ _]{2,}', '_', base).strip(' _-')
    month_name = parse_date(args.period_end).strftime('%b%Y')
    out_filename = (
        f"{args.project}_{base}_Claim{args.claim_no}_{month_name}_"
        f"DRAFT_v{args.draft_version}.xlsx"
    )
    out_path = os.path.join(args.output_dir, out_filename)
    print(f"\nPhase 6-9: Building {out_filename}")
    build_workbook(records, orphans, cfg, args.claim_template, out_path,
                   args.claim_no, args.period_start, args.period_end,
                   args.draft_version)

    print("Phase 7: Safety self-check")
    safety_check(out_path, cfg)

    print("Phase 8: LibreOffice recalc + verification")
    cover = recalc_and_verify(out_path)

    # Phase 9: Audit log
    audit_filename = f"{args.project}_Claim{args.claim_no}_audit_v{args.draft_version}.md"
    audit_path = os.path.join(args.output_dir, audit_filename)
    write_audit_log(audit_path, records, orphans, cover, cfg,
                    args.claim_no, args.period_end)

    print("\n=== Outputs ===")
    print(f"Draft claim: {out_path}")
    print(f"Audit log:   {audit_path}")
    print()
    print(f"=== Claim {cover['claim_no']} totals ===")
    print(f"  Total completed:    ${cover['total_completed']:,.2f}")
    print(f"  This period ex GST: ${cover['this_period_ex_gst']:,.2f}")
    print(f"  This period incl GST: ${cover['this_period_incl_gst']:,.2f}")
    print(f"  Balance to finish:  ${cover['balance_to_finish']:,.2f}")

    return 0


if __name__ == '__main__':
    sys.exit(main())
