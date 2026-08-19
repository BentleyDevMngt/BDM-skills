#!/usr/bin/env python3
"""
build_certificate.py — BDM Progress Certificate (Form 335) builder.

Takes the latest 335 template + a project/claim config (JSON) and produces:
  - a populated, macro-free .xlsm (Excel recalculates on open)
  - a DRAFT PDF of tabs 01-04 only (Cover, Certificate, Trade Breakdown, Cashflow)
and verifies the net-this-claim incl GST equals the builder's tax invoice.

Design split: Claude reads the project folder + claim docs and assembles the
config (judgement); this script does the deterministic build + verification.

Staging is operator-agnostic: pass --outdir pointing at the active project's
00_ai_sandbox\\PC 0NN (Month YYYY) folder. The script writes ONLY the final
.xlsm and .pdf there; all of its own intermediates live in a private temp dir
that is always removed on exit.

Dependencies: python3, openpyxl, and LibreOffice (`soffice`) on PATH.

Usage:
  python build_certificate.py --template <335_R*.xlsm> --config <config.json> --outdir <dir>

The guard functions are idempotent: they fix the known R3 template defects but
are no-ops on a clean R4+ master, so this script works on either.
"""
import argparse, json, math, os, re, shutil, subprocess, sys, tempfile, datetime
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.workbook.properties import CalcProperties

KEEP = ['01 Cover Letter', '02 Certificate', '03 Trade Breakdown', '04 Cashflow Forecast']
DROP = ['00 Project Details', 'SINGLE STAGE (3)', 'CashflowInput', 'Data']
EXCEL_ERRORS = ('#REF!', '#DIV/0!', '#VALUE!', '#NAME?', '#N/A', '#NUM!', '#NULL!')


def _date(s):
    if isinstance(s, datetime.datetime):
        return s
    return datetime.datetime.strptime(s, "%Y-%m-%d")


# ---------------------------------------------------------------- template guards
def stamp_revision(wb, rev_label):
    """Refresh stale print header/footer revision stamps on every sheet."""
    rev_no = rev_label.split()[0]
    rev_date = rev_label.split()[-1]
    for ws in wb.worksheets:
        for part in (ws.oddHeader, ws.oddFooter):
            for side in (part.left, part.center, part.right):
                t = side.text
                if not t:
                    continue
                t = re.sub(r'R\d+', rev_no, t)
                t = re.sub(r'20\d\d-\d\d', rev_date, t)
                side.text = t
    if "Form 335" in str(wb['01 Cover Letter']['C1'].value or ""):
        wb['01 Cover Letter']['C1'] = f"Form 335  ·  {rev_no}  ·  {rev_date}"


def fix_cover_certno(wb):
    """Cover letter must read the certificate NUMBER from 02!F6, not the title cell A4."""
    cl = wb['01 Cover Letter']
    for coord in ('B17', 'B19', 'B22'):
        v = cl[coord].value
        if isinstance(v, str) and "'02 Certificate'!A4" in v:
            cl[coord] = v.replace("'02 Certificate'!A4", "'02 Certificate'!F6")


def port_scurve(wb):
    """Replace the VBA scurve() UDF with native Excel formulas (helper cols J/K on
    CashflowInput, outside the A14:H54 print area). Macro-free => prints everywhere."""
    ci = wb['CashflowInput']
    rows = sorted({c.row for row in ci.iter_rows() for c in row
                   if isinstance(c.value, str) and 'scurve' in c.value})
    for r in rows:
        ci[f"J{r}"] = f'=IF($B{r}="","",($B{r}-$D$8)/($D$10-$D$8))'
        ci[f"K{r}"] = (f'=IF($J{r}="","",$J{r}-0.016*$J{r}^2+0.016*$J{r}'
                       f'-1/3.021*(6*$J{r}^3-9*$J{r}^2+3*$J{r}))')
        ci[f"F{r}"] = (f'=IF(OR($A{r}="",$B{r}="",$D$7=""),0,'
                       f'IF($J{r}=1,$D$7*$K{r},INT($D$7*$K{r}/1000)*1000))')
    if rows:
        ci["J28"], ci["K28"] = "x(h)", "y(h)"
    return len(rows)


def fix_cashflow_cumulative(wb):
    """04 'Actual Cumulative' (E) total should take the last value, not SUM."""
    cf = wb['04 Cashflow Forecast']
    e33 = cf['E33'].value
    if isinstance(e33, str) and 'SUM' in e33:
        cf['E33'] = '=MAX(E12:E32)'


def remove_input_highlight(wb):
    """Clear the yellow 'INPUT DATES' highlight on the cashflow note cells."""
    cf = wb['04 Cashflow Forecast']
    nofill = PatternFill(fill_type=None)
    for r in (75, 76):
        for c in (4, 5, 6):
            cf.cell(r, c).fill = nofill


def apply_guards(wb, rev_label):
    fix_cover_certno(wb)
    n = port_scurve(wb)
    fix_cashflow_cumulative(wb)
    remove_input_highlight(wb)
    stamp_revision(wb, rev_label)
    return n


# ---------------------------------------------------------------- population
def populate(wb, cfg):
    d = wb['00 Project Details']
    pd = cfg['project']
    fields = {
        'C8': pd['name'], 'C9': pd['address'], 'C10': pd['principal'],
        'C11': pd['contractor'], 'C12': pd.get('client', pd['principal']),
        'C13': pd.get('development_type', ''), 'C14': pd.get('stage', 'Construction'),
        'C15': cfg['contract_sum'], 'C19': cfg.get('revision', 'DRAFT'),
        'C20': 'Draft', 'C21': cfg.get('signatory', '[For signature — Senior QS / Director]'),
        'C22': cfg.get('reviewer', '[Reviewer]'),
        'C25': pd.get('builder_contact', ''), 'C26': pd.get('builder_address', ''),
        'C27': pd.get('builder_email', ''),
    }
    for k, v in fields.items():
        d[k] = v
    d['C18'] = _date(cfg['issue_date'])

    tb = wb['03 Trade Breakdown']
    for i, tr in enumerate(cfg['trades']):
        r = 14 + i
        tb.cell(r, 1, tr['desc'])
        tb.cell(r, 3, tr['original'])
        tb.cell(r, 7, tr.get('prev', 0))
        tb.cell(r, 9, tr.get('todate', 0))
        tb.row_dimensions[r].height = 30
    tb.column_dimensions['A'].width = 34

    ret = cfg['retention']
    rate = ret['rate']
    if ret.get('cap_pct'):
        cap = ret['cap_pct']
        tb['I76'] = f"=-MIN((I46+I71)*{rate},{cfg['contract_sum']}*{cap})"
        tb['A76'] = (f"Less Cash Retention ({rate*100:g}% of value to date, "
                     f"capped at {cap*100:g}% of Contract Sum — per Contract)")
    else:
        tb['I76'] = f"=-(I46+I71)*{rate}"
        tb['A76'] = f"Less Cash Retention ({rate*100:g}% — per Contract)"

    c = wb['02 Certificate']
    c['F6'] = cfg['cert_no']
    c['F5'] = _date(cfg['valuation_date'])
    c['F32'] = -abs(cfg['previous_net'])

    cf_form = cfg.get('contract_form', 'AS4000-2024')
    cl = wb['01 Cover Letter']
    cl['B19'] = ('="Pursuant to cl.37.2 of the General Conditions of the ' + cf_form +
                 ' Contract, Bentley Development Management issues Progress Certificate No. "'
                 '&TEXT(\'02 Certificate\'!F6,"00")&" for "&\'00 Project Details\'!C11'
                 '&" relating to "&\'00 Project Details\'!C8&"."')
    cl['B39'] = 'BENTLEY DEVELOPMENT MANAGEMENT  ·  SUPERINTENDENT UNDER ' + cf_form.upper()
    cl['B42'] = ('This Certificate is issued under ' + cf_form + ' cl.37.2. The Superintendent '
                 'is required to issue this Progress Certificate within the time stated in the '
                 'Contract of receipt of the Claim.')
    c['B59'] = ('Issued under ' + cf_form + " cl.37.2. Superintendent must issue within the time "
                "stated in the Contract of receipt of the Builder's Progress Claim.")

    cfo = cfg.get('cashflow', {})
    ci = wb['CashflowInput']
    ci['D4'] = pd.get('project_code', 0)
    ci['D7'] = cfo.get('total', cfg['contract_sum'])
    if cfo.get('start'):
        ci['D8'] = _date(cfo['start'])
    if cfo.get('finish'):
        ci['D10'] = _date(cfo['finish'])

    cf = wb['04 Cashflow Forecast']
    actuals = {a['month']: a['monthly'] for a in cfo.get('actuals', [])}
    for r in range(12, 33):
        cf.cell(r, 3).value = None
        cf.cell(r, 5).value = None
        cf.cell(r, 6).value = f'=IF($E{r}="","",$E{r}-$D{r})'
    cum = 0
    for i in range(13):
        r = 12 + i
        bd = ci.cell(30 + i, 2).value
        if not isinstance(bd, datetime.datetime):
            continue
        ym = bd.strftime('%Y-%m')
        if ym in actuals:
            cum += actuals[ym]
            cf.cell(r, 3, actuals[ym])
            cf.cell(r, 5, cum)
    cf['E33'] = '=MAX(E12:E32)'


# ---------------------------------------------------------------- verification
def verify_python(cfg):
    I46 = sum(t.get('todate', 0) for t in cfg['trades'])
    G46 = sum(t.get('prev', 0) for t in cfg['trades'])
    ret = cfg['retention']
    base = I46 * ret['rate']
    if ret.get('cap_pct'):
        base = min(base, cfg['contract_sum'] * ret['cap_pct'])
    retention = -base
    net_to_date = I46 + retention
    net_this = net_to_date - abs(cfg['previous_net'])
    incl = round(net_this * 1.1, 2)
    out = {'claimed_to_date': I46, 'prev_claimed': G46, 'this_claim': I46 - G46,
           'retention': retention, 'net_to_date': net_to_date,
           'net_this_claim_excl': net_this, 'net_this_claim_incl': incl}
    inv = cfg.get('invoice_total')
    out['invoice_total'] = inv
    out['ties_to_invoice'] = (inv is None) or abs(incl - inv) < 0.5
    return out


def _soffice(args):
    env = dict(os.environ, HOME=tempfile.gettempdir())
    subprocess.run(['soffice', '--headless', '--calc'] + args,
                   check=True, env=env,
                   stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)


def recalc_and_pdf(xlsm_path, pdf_path):
    """Recalc via LibreOffice, export tabs 01-04 only to PDF (static-ize + delete the rest).
    All intermediates live in a private temp dir that is ALWAYS removed on exit."""
    tmp = tempfile.mkdtemp()
    try:
        _soffice(['--convert-to', 'xlsx', '--outdir', tmp, xlsm_path])
        recalced = os.path.join(tmp, os.path.splitext(os.path.basename(xlsm_path))[0] + '.xlsx')
        vals = openpyxl.load_workbook(recalced, data_only=True)
        errors = []
        for ws in vals.worksheets:
            for row in ws.iter_rows():
                for c in row:
                    if isinstance(c.value, str) and c.value in EXCEL_ERRORS:
                        errors.append(f"{ws.title}!{c.coordinate}={c.value}")
        wb = openpyxl.load_workbook(recalced)
        for name in KEEP:
            ws, vs = wb[name], vals[name]
            for row in ws.iter_rows():
                for c in row:
                    if isinstance(c.value, str) and c.value.startswith('='):
                        c.value = vs[c.coordinate].value
        for name in DROP:
            if name in wb.sheetnames:
                del wb[name]
        wb.active = 0
        core = os.path.join(tmp, 'core.xlsx')
        wb.save(core)
        _soffice(['--convert-to', 'pdf', '--outdir', tmp, core])
        shutil.move(os.path.join(tmp, 'core.pdf'), pdf_path)
        return errors
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


# ---------------------------------------------------------------- main
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--template', required=True)
    ap.add_argument('--config', required=True)
    ap.add_argument('--outdir', required=True,
                    help="staging dir — the project's 00_ai_sandbox\\PC 0NN (Month YYYY)")
    ap.add_argument('--rev', default='R4 2026-06', help='revision label to stamp')
    a = ap.parse_args()

    cfg = json.load(open(a.config, encoding='utf-8'))
    os.makedirs(a.outdir, exist_ok=True)
    base = cfg['output_basename']
    xlsm = os.path.join(a.outdir, base + '.xlsm')
    pdf = os.path.join(a.outdir, base + '.pdf')

    wb = openpyxl.load_workbook(a.template, keep_vba=True)
    ported = apply_guards(wb, a.rev)
    populate(wb, cfg)
    wb.calculation = CalcProperties(fullCalcOnLoad=True)
    wb.save(xlsm)

    chk = verify_python(cfg)
    errors = recalc_and_pdf(xlsm, pdf)

    print(json.dumps({
        'xlsm': xlsm, 'pdf': pdf, 'scurve_rows_ported': ported,
        'formula_errors': errors[:20], 'error_count': len(errors),
        'verification': chk,
        'PASS': chk['ties_to_invoice'] and not errors,
    }, indent=2, default=str))
    if errors or not chk['ties_to_invoice']:
        sys.exit(1)


if __name__ == '__main__':
    main()
