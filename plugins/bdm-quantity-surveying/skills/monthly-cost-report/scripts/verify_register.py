#!/usr/bin/env python3
"""Pre-delivery check for a Form 201 consultant fee register.

Checks the things that have actually gone wrong in practice, each of which produces a
file that opens cleanly and is wrong:

  * block subtotals not reconciling to the grand total (a Summary row pointing at the
    wrong Breakdown row reads $0 and nothing complains)
  * the Fees Invoiced columns not adding up
  * print settings lost against the base file the user had tuned
  * the header logo stripped, or its relationships broken
  * cell comments destroyed by a logo repair
  * leftover bracketed status text in the Stage column
  * a stale xl/calcChain.xml, which makes Excel report the file as corrupt even though
    openpyxl and LibreOffice both read it happily (delegated to validate_xlsx.py)

Usage:
    # structure, print settings and the header logo — always run on the DELIVERABLE
    python verify_register.py <output.xlsx> --base <base.xlsx>

    # value checks need calculated formulas, so recalculate to a throwaway copy first:
    soffice --headless --convert-to xlsx --outdir rc <output.xlsx>
    python verify_register.py <output.xlsx> --base <base.xlsx> --calc rc/<output.xlsx>

Never run the structure or logo checks against the recalculated copy — LibreOffice
rewrites the print scale and drops the header image, so it fails checks the real file
passes. That is the same trap described in references/register-mechanics.md.

Exit code is 0 if every check passes, 1 otherwise, so it can gate delivery.
"""
import argparse
import os
import re
import sys
import zipfile
from xml.etree import ElementTree as ET

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl required: pip install openpyxl --break-system-packages")

RELNS = '{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'
SUMMARY = 'Professional Fees Summary'
BREAKDOWN = 'Professional Fees Breakdown'

results = []


def check(name, ok, detail=''):
    results.append((name, bool(ok), detail))
    print(f"[{'PASS' if ok else 'FAIL'}] {name}" + (f" — {detail}" if detail else ''))
    return ok


def money(v):
    return round(v or 0, 2)


def find_summary_rows(ws):
    """Rows 8..N of the Summary that carry a discipline reference, plus the TOTAL row."""
    blocks, total_row = [], None
    for r in range(8, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if isinstance(a, str) and a.strip().upper() == 'TOTAL':
            total_row = r
        elif a not in (None, ''):
            blocks.append(r)
    return blocks, total_row


def verify_values(path, expect_total=None):
    """Value checks need a workbook whose formulas have been calculated."""
    wb = openpyxl.load_workbook(path, data_only=True)
    if SUMMARY not in wb.sheetnames:
        return check('summary sheet present', False, f'sheets: {wb.sheetnames}')
    sm = wb[SUMMARY]
    blocks, total_row = find_summary_rows(sm)
    if not blocks or not total_row:
        print('  ! the Summary reads back as formulas, so values cannot be checked here.')
        print('    Recalculate first, then pass the calculated copy via --calc:')
        print('      soffice --headless --convert-to xlsx --outdir rc <file>')
        return check('summary values checkable', False,
                     'no calculated copy supplied (--calc)')

    vals = [sm.cell(total_row, c).value for c in (3, 4, 5, 7, 8, 9, 10)]
    if any(isinstance(v, str) for v in vals):
        print('  ! formulas not calculated — recalculate before trusting value checks:')
        print('    soffice --headless --convert-to xlsx --outdir rc <file>')
        return check('workbook is calculated', False, 'formula strings returned')

    cols = {c: money(sum(sm.cell(r, c).value or 0 for r in blocks))
            for c in (3, 4, 5, 7, 8, 9, 10)}
    tot = {c: money(sm.cell(total_row, c).value) for c in (3, 4, 5, 7, 8, 9, 10)}

    check('block subtotals reconcile to grand total',
          cols == tot,
          '' if cols == tot else f'blocks {cols} vs total {tot}')
    check('Total = Approved + Variations',
          money(tot[3] + tot[4]) == tot[5],
          f'{tot[3]} + {tot[4]} != {tot[5]}' if money(tot[3] + tot[4]) != tot[5] else '')
    check('This month = To date - Previous',
          money(tot[7] - tot[8]) == tot[9],
          f'{tot[7]} - {tot[8]} != {tot[9]}' if money(tot[7] - tot[8]) != tot[9] else '')
    check('Remaining = Total - Invoiced to date',
          money(tot[5] - tot[7]) == tot[10],
          f'{tot[5]} - {tot[7]} != {tot[10]}' if money(tot[5] - tot[7]) != tot[10] else '')
    if expect_total is not None:
        check(f'grand total equals expected {expect_total:,.2f}',
              money(expect_total) == tot[5], f'got {tot[5]:,.2f}')
    print(f"\n  Approved {tot[3]:,.2f} | Variations {tot[4]:,.2f} | Total {tot[5]:,.2f}"
          f" | Invoiced {tot[7]:,.2f} (prev {tot[8]:,.2f}, this month {tot[9]:,.2f})"
          f" | Remaining {tot[10]:,.2f}\n")
    return True


def verify_structure(path, base=None):
    wb = openpyxl.load_workbook(path)
    bd = wb[BREAKDOWN] if BREAKDOWN in wb.sheetnames else None

    if bd is not None:
        stray = [r for r in range(9, bd.max_row + 1)
                 if isinstance(bd.cell(r, 3).value, str) and '[' in bd.cell(r, 3).value]
        check('no bracketed status text left in the Stage column', not stray,
              f'rows {stray[:8]}' if stray else 'status belongs in cell comments')

        n = sum(1 for row in bd.iter_rows(min_row=9, max_col=3) for c in row if c.comment)
        check('cell comments present', n > 0, f'{n} comments')

        pa = bd.print_area
        pa = pa[0] if isinstance(pa, (list, tuple)) and pa else pa
        last_content = max((r for r in range(1, bd.max_row + 1)
                            if any(bd.cell(r, c).value not in (None, '') for c in range(1, 12))),
                           default=0)
        m = re.search(r'\$(\d+)$', pa or '')
        check('print area covers all content',
              bool(m) and int(m.group(1)) >= last_content,
              f'area to {m.group(1) if m else "?"}, content to {last_content}')

    if base:
        bw = openpyxl.load_workbook(base)
        for name in wb.sheetnames:
            if name not in bw.sheetnames:
                continue
            a, b = wb[name], bw[name]
            same = (a.page_setup.scale == b.page_setup.scale
                    and a.page_setup.orientation == b.page_setup.orientation
                    and a.page_setup.fitToHeight == b.page_setup.fitToHeight
                    and a.page_setup.fitToWidth == b.page_setup.fitToWidth
                    and a.print_title_rows == b.print_title_rows
                    and a.print_options.horizontalCentered == b.print_options.horizontalCentered
                    and round(a.page_margins.left or 0, 4) == round(b.page_margins.left or 0, 4)
                    and round(a.page_margins.top or 0, 4) == round(b.page_margins.top or 0, 4))
            check(f"print settings preserved — {name}", same,
                  f'scale={a.page_setup.scale} titles={a.print_title_rows}')


def verify_parts(path, base=None):
    """The header logo and its relationships. openpyxl strips these on every save."""
    z = zipfile.ZipFile(path)
    names = set(z.namelist())

    check('header image present', any(n.startswith('xl/media/') for n in names),
          'run fix_header_logo.py after every save' if not any(
              n.startswith('xl/media/') for n in names) else '')

    ok, seen = True, 0
    for sheet in sorted(n for n in names if re.match(r'xl/worksheets/sheet\d+\.xml$', n)):
        i = re.search(r'sheet(\d+)\.xml', sheet).group(1)
        relp = f'xl/worksheets/_rels/sheet{i}.xml.rels'
        rmap = ({r.get('Id'): r.get('Target') for r in ET.fromstring(z.read(relp)).iter(RELNS)}
                if relp in names else {})
        s = z.read(sheet).decode('utf8')
        for tag in ('legacyDrawingHF', 'legacyDrawing'):
            for m in re.finditer(r'<' + tag + r'\s[^>]*r:id="([^"]+)"', s):
                seen += 1
                tgt = rmap.get(m.group(1), '')
                part = tgt.lstrip('/') if tgt.startswith('/') else 'xl/' + tgt.replace('../', '')
                if part not in names:
                    ok = False
                    print(f'    broken: sheet{i} {tag} -> {tgt or "(no relationship)"}')
        if '&amp;G' in s and 'legacyDrawingHF' not in s:
            ok = False
            print(f'    sheet{i}: header requests a graphic (&G) but has no legacyDrawingHF')
    check('drawing relationships resolve', ok, f'{seen} references checked')

    if base:
        zb = zipfile.ZipFile(base)
        imgs = [n for n in names if n.startswith('xl/media/')]
        bimgs = [n for n in zb.namelist() if n.startswith('xl/media/')]
        if imgs and bimgs:
            import hashlib
            h = {hashlib.sha256(z.read(n)).hexdigest() for n in imgs}
            hb = {hashlib.sha256(zb.read(n)).hexdigest() for n in bimgs}
            check('logo image identical to source', bool(h & hb))


def verify_excel_integrity(path):
    """Excel-strict package checks — the ones openpyxl and LibreOffice cannot see.

    Chiefly a stale xl/calcChain.xml: it indexes every formula cell, and rolling the
    register forward turns formula cells into values, so the index ends up naming cells
    with no formula. Excel walks it on open and calls the file corrupt. Fix with
    `python validate_xlsx.py --repair <file>`.
    """
    try:
        sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
        from validate_xlsx import check as _xlsx_check
    except ImportError:
        check('Excel package integrity', False,
              'validate_xlsx.py not found alongside this script — cannot check calcChain')
        return
    errs = _xlsx_check(path)
    check('Excel package integrity (calcChain, dimension, namespaces)', not errs,
          '' if not errs else f'{len(errs)} defect(s); run validate_xlsx.py --repair')
    for e in errs:
        print(f'    {e}')


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('workbook')
    ap.add_argument('--base', help='the file this was built from, to compare print settings')
    ap.add_argument('--calc', help='a LibreOffice-recalculated copy, for the value checks')
    ap.add_argument('--expect-total', type=float)
    a = ap.parse_args()

    print(f'Verifying {a.workbook}\n')
    verify_values(a.calc or a.workbook, a.expect_total)
    verify_structure(a.workbook, a.base)
    verify_parts(a.workbook, a.base)
    verify_excel_integrity(a.workbook)

    failed = [n for n, ok, _ in results if not ok]
    print(f"\n{len(results) - len(failed)}/{len(results)} checks passed")
    if failed:
        print('Failed: ' + '; '.join(failed))
    return 1 if failed else 0


if __name__ == '__main__':
    sys.exit(main())
