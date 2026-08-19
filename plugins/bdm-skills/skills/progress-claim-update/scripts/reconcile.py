#!/usr/bin/env python3
"""
reconcile.py — R4 hard gate between the CTD Breakdown and the Xero export.

Why this exists
---------------
Every internal check in the R3 pipeline passed on the PC18 (July 2026) draft, and
that draft still overstated the month by $84,476.50. The checks proved the workbook
was self-consistent; nothing proved it matched the bank.

This module answers one question: does the CTD Breakdown, for the claim period,
contain exactly the transactions the Xero account-transactions export contains —
once each? It reconciles on (date, amount) so it does not depend on supplier-name
parsing, which is the very thing that fails.

Run it standalone at any time:
    python3 reconcile.py --claim CLAIM.xlsx --xero XERO.xlsx \
        --period-start 24/06/2026 --period-end 27/07/2026 [--prior PRIOR_CLAIM.xlsx]

Exit code 0 = reconciled (or only accepted exceptions), 1 = variance requiring review.
"""
from __future__ import annotations
import argparse, datetime, re, sys
from collections import defaultdict, Counter
from openpyxl import load_workbook

TOL = 0.02


# ---------------------------------------------------------------- loading
def _subtotal_rows(ws_f):
    return {r for r in range(1, ws_f.max_row + 1)
            if isinstance(ws_f.cell(r, 5).value, str)
            and ws_f.cell(r, 5).value.startswith('=SUBTOTAL')}


def load_ctd(claim_path, date_from=None, date_to=None):
    """CTD detail rows as (row, date, description, gross). Subtotal rows excluded."""
    wv = load_workbook(claim_path, data_only=True)
    wf = load_workbook(claim_path, data_only=False)
    cs, cf = wv['CTD Breakdown'], wf['CTD Breakdown']
    subs = _subtotal_rows(cf)
    out = []
    for r in range(7, cf.max_row + 1):
        if r in subs:
            continue
        d = cs.cell(r, 2).value
        if not isinstance(d, datetime.datetime):
            continue
        if date_from and d < date_from:
            continue
        if date_to and d > date_to:
            continue
        g = cs.cell(r, 7).value
        g = float(g) if isinstance(g, (int, float)) else 0.0
        desc = str(cs.cell(r, 4).value or '')
        if desc.startswith('[REMOVED'):        # zeroed and marked — not a live row
            continue
        out.append({'row': r, 'date': d, 'desc': desc, 'gross': round(g, 2)})
    return out


def load_xero(xero_path, sheet='Construction Costs Transactions'):
    ws = load_workbook(xero_path, data_only=True)[sheet]
    out = []
    for r in ws.iter_rows(min_row=7, values_only=True):
        d = r[0]
        if not isinstance(d, datetime.datetime):
            continue
        out.append({'date': d, 'desc': str(r[2] or '').strip(),
                    'gross': round(float(r[7] or 0), 2)})
    return out


# ---------------------------------------------------------------- matching
def reconcile(ctd_rows, xero_rows):
    """Multiset match on (date, gross), then a looser amount-only pass.

    Returns dict with matched / ctd_only / xero_only / date_mismatch.
    """
    x_pool = defaultdict(list)
    for t in xero_rows:
        x_pool[(t['date'].date(), t['gross'])].append(t)

    matched, ctd_only = [], []
    for row in ctd_rows:
        key = (row['date'].date(), row['gross'])
        if x_pool.get(key):
            matched.append((row, x_pool[key].pop(0)))
        else:
            ctd_only.append(row)

    leftover_x = [t for lst in x_pool.values() for t in lst]

    # second pass: same amount, different date (a re-dated but genuine entry)
    amt_pool = defaultdict(list)
    for t in leftover_x:
        amt_pool[t['gross']].append(t)
    date_mismatch, still_ctd_only = [], []
    for row in ctd_only:
        cand = amt_pool.get(row['gross'])
        if cand:
            date_mismatch.append((row, cand.pop(0)))
        else:
            still_ctd_only.append(row)
    xero_only = [t for lst in amt_pool.values() for t in lst]

    return {'matched': matched, 'date_mismatch': date_mismatch,
            'ctd_only': still_ctd_only, 'xero_only': xero_only}


def summarise(res, ctd_rows, xero_rows):
    ctd_total = round(sum(r['gross'] for r in ctd_rows), 2)
    xero_total = round(sum(t['gross'] for t in xero_rows), 2)
    return {
        'ctd_rows': len(ctd_rows), 'xero_rows': len(xero_rows),
        'ctd_total': ctd_total, 'xero_total': xero_total,
        'variance': round(ctd_total - xero_total, 2),
        'matched': len(res['matched']),
        'date_mismatch': len(res['date_mismatch']),
        'ctd_only': len(res['ctd_only']),
        'xero_only': len(res['xero_only']),
        'ctd_only_value': round(sum(r['gross'] for r in res['ctd_only']), 2),
        'xero_only_value': round(sum(t['gross'] for t in res['xero_only']), 2),
    }


# ---------------------------------------------------------------- extra checks
def prior_claim_duplicates(claim_path, prior_path, date_from):
    """Rows in this claim's period that already appear in the prior claim's CTD.

    PC18 caught SiteSec $1,061.50 and Superior Steel $426.00 this way; the R3
    pipeline only de-duplicated within the current month.
    """
    prior = load_ctd(prior_path)
    prior_keys = Counter((p['date'].date(), p['gross']) for p in prior)
    hits = []
    for row in load_ctd(claim_path, date_from=date_from):
        k = (row['date'].date(), row['gross'])
        if prior_keys.get(k):
            hits.append(row)
    return hits


def cover_wiring(claim_path):
    """The cover's 'less previous progress claims' must be the prior claim's
    total, i.e. Detail Sheet H80. Two different PC18 drafts had this wrong in two
    different ways — hardcoded to cash received, and pointed at I80 (this period).
    """
    wv = load_workbook(claim_path, data_only=True)
    wf = load_workbook(claim_path, data_only=False)
    pc_v, pc_f, ds = wv['Progress Claim'], wf['Progress Claim'], wv['Detail Sheet']
    f28_formula = pc_f['F28'].value
    issues = []
    if not isinstance(f28_formula, str) or 'H80' not in f28_formula.replace(' ', ''):
        issues.append(f"cover F28 is {f28_formula!r}; expected =\'Detail Sheet\'!H80")
    if abs((pc_v['F28'].value or 0) - (ds['H80'].value or 0)) > TOL:
        issues.append(f"cover F28 {pc_v['F28'].value:,.2f} != Detail Sheet H80 {ds['H80'].value:,.2f}")
    if abs((pc_v['F22'].value or 0) - (ds['J80'].value or 0)) > TOL:
        issues.append("cover F22 != Detail Sheet J80")
    if abs((pc_v['F31'].value or 0) - ((pc_v['F29'].value or 0) + (pc_v['F30'].value or 0))) > TOL:
        issues.append("cover F31 != F29 + F30")
    return issues


def manual_journals(claim_path):
    """CTD rows whose Xero source is a journal, not a payment. These belong in
    cost-to-date but not in a payments figure, and they are never reversed in the
    CTD — Everton carried $143,792.49 of them undetected for a year.
    """
    wv = load_workbook(claim_path, data_only=True)
    wf = load_workbook(claim_path, data_only=False)
    cs, cf = wv['CTD Breakdown'], wf['CTD Breakdown']
    subs = _subtotal_rows(cf)
    out = []
    for r in range(7, cf.max_row + 1):
        if r in subs:
            continue
        if str(cs.cell(r, 3).value or '').strip().lower() == 'manual journal':
            out.append({'row': r, 'date': cs.cell(r, 2).value,
                        'desc': str(cs.cell(r, 4).value or ''),
                        'debit': float(cs.cell(r, 5).value or 0)})
    return out


# ---------------------------------------------------------------- CLI
def _d(s):
    return datetime.datetime.strptime(s, '%d/%m/%Y')


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--claim', required=True)
    ap.add_argument('--xero', required=True)
    ap.add_argument('--period-start', required=True)
    ap.add_argument('--period-end', required=True)
    ap.add_argument('--prior')
    ap.add_argument('--accept', type=float, default=0.0,
                    help='value of Director-accepted out-of-period items (e.g. a paid '
                         'invoice dated after period end) to net off the variance')
    a = ap.parse_args()
    ds, de = _d(a.period_start), _d(a.period_end)

    xero = load_xero(a.xero)
    xds = min(t['date'] for t in xero); xde = max(t['date'] for t in xero)
    ctd = load_ctd(a.claim, date_from=xds, date_to=max(de, xde))

    # Rows already carried in the prior claim are legitimately in the CTD but were
    # claimed in an earlier period; they are not part of this period's Xero set.
    carried = []
    if a.prior:
        prior_keys = Counter((p['date'].date(), p['gross']) for p in load_ctd(a.prior))
        keep = []
        for row in ctd:
            k = (row['date'].date(), row['gross'])
            if prior_keys.get(k):
                prior_keys[k] -= 1
                carried.append(row)
            else:
                keep.append(row)
        ctd = keep

    # Rows dated after the Xero export ends cannot be reconciled against it.
    post = [r for r in ctd if r['date'] > xde]
    ctd = [r for r in ctd if r['date'] <= xde]

    res = reconcile(ctd, xero)
    s = summarise(res, ctd, xero)

    print('=' * 78)
    print(f"CTD <-> XERO RECONCILIATION   period {a.period_start} to {a.period_end}")
    print('=' * 78)
    print(f"  Xero transactions {s['xero_rows']:>5}   total {s['xero_total']:>15,.2f}")
    print(f"  CTD rows          {s['ctd_rows']:>5}   total {s['ctd_total']:>15,.2f}")
    print(f"  VARIANCE                        {s['variance']:>15,.2f}")
    print(f"  matched {s['matched']}   date-only mismatches {s['date_mismatch']}"
          f"   CTD-only {s['ctd_only']}   Xero-only {s['xero_only']}")

    if res['ctd_only']:
        print(f"\n  IN THE CLAIM BUT NOT IN XERO  ({s['ctd_only_value']:,.2f}) "
              f"— duplicates or unsupported costs:")
        for r in sorted(res['ctd_only'], key=lambda z: -abs(z['gross'])):
            print(f"    r{r['row']:<5} {r['date']:%d/%m/%Y} {r['gross']:>13,.2f}  {r['desc'][:62]}")
    if res['xero_only']:
        print(f"\n  PAID IN XERO BUT NOT IN THE CLAIM  ({s['xero_only_value']:,.2f}) — omissions:")
        for t in sorted(res['xero_only'], key=lambda z: -abs(z['gross'])):
            print(f"          {t['date']:%d/%m/%Y} {t['gross']:>13,.2f}  {t['desc'][:62]}")
    if carried:
        print(f"\n  CARRIED FROM THE PRIOR CLAIM (already claimed, excluded from the variance) "
              f"— {sum(c['gross'] for c in carried):,.2f}:")
        for c in carried:
            print(f"    r{c['row']:<5} {c['date']:%d/%m/%Y} {c['gross']:>13,.2f}  {c['desc'][:62]}")
    if res['date_mismatch']:
        print("\n  RIGHT AMOUNT, WRONG DATE (no effect on the total — invoice date vs payment date):")
        for r, t in res['date_mismatch']:
            print(f"    r{r['row']:<5} claim {r['date']:%d/%m/%Y} vs Xero {t['date']:%d/%m/%Y}"
                  f" {r['gross']:>13,.2f}  {t['desc'][:52]}")

    if post:
        print(f"\n  DATED AFTER THE XERO EXPORT ENDS ({xde:%d/%m/%Y}) — cannot be reconciled "
              f"against it; confirm with the Director — {sum(p['gross'] for p in post):,.2f}:")
        for r in post:
            print(f"    r{r['row']:<5} {r['date']:%d/%m/%Y} {r['gross']:>13,.2f}  {r['desc'][:62]}")

    issues = cover_wiring(a.claim)
    print("\n  COVER WIRING:", 'OK' if not issues else '')
    for i in issues:
        print("    FAIL:", i)

    mj = manual_journals(a.claim)
    if mj:
        print(f"\n  MANUAL JOURNALS IN THE CTD (cost, not cash) — {sum(m['debit'] for m in mj):,.2f}:")
        for m in mj:
            print(f"    r{m['row']:<5} {m['date']:%d/%m/%Y} {m['debit']:>13,.2f}  {m['desc'][:62]}")

    net = round(s['variance'] - a.accept, 2)
    ok = abs(net) <= TOL and not issues
    print('\n' + '=' * 78)
    print(f"  VERDICT: {'RECONCILED' if ok else 'VARIANCE — DO NOT ISSUE'}"
          f"   (net of accepted {a.accept:,.2f}: {net:,.2f})")
    print('=' * 78)
    return 0 if ok else 1


if __name__ == '__main__':
    sys.exit(main())
