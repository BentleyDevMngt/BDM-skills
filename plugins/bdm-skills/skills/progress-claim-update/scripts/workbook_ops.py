#!/usr/bin/env python3
"""
workbook_ops.py — R4 safe mechanics for editing a live Progress Claim workbook.

Each function here exists because the corresponding operation went wrong on a real
claim. Read the docstrings before changing any of them.
"""
from __future__ import annotations
import re
from copy import copy
from openpyxl.styles import Font, PatternFill

REMOVED_PREFIX = '[REMOVED'
GREY = PatternFill('solid', fgColor='FFEFEFEF')
RED = Font(name='Calibri', size=11, color='FFC00000')


def subtotal_rows(ws_f):
    return [r for r in range(1, ws_f.max_row + 1)
            if isinstance(ws_f.cell(r, 5).value, str)
            and ws_f.cell(r, 5).value.startswith('=SUBTOTAL')]


def is_grand_total_row(ws_f, r) -> bool:
    """The sheet-wide Total spans the whole CTD, not just the last block.

    R3's rule — 'each subtotal range runs from the previous subtotal + 1 to this
    row - 1' — is correct for supplier blocks and WRONG here. Applying it to the
    grand total silently reduced =SUBTOTAL(9,E7:E908) to =SUBTOTAL(9,E911:E912),
    i.e. zero, on the PC18 v3.0 draft. Nothing downstream reads the grand total,
    so every other check still passed; it was only caught when the Director asked
    what the CTD totalled.
    """
    label = str(ws_f.cell(r, 2).value or ws_f.cell(r, 4).value or '').strip().lower()
    return label in ('total', 'grand total', 'total ')


def rebuild_subtotals(ws_f, first_data_row: int = 7):
    """Recompute every block SUBTOTAL range; span the whole sheet for the grand total."""
    subs = subtotal_rows(ws_f)
    if not subs:
        return 0
    last_block = max((r for r in subs if not is_grand_total_row(ws_f, r)), default=first_data_row)
    prev, n = first_data_row - 1, 0
    for r in subs:
        if is_grand_total_row(ws_f, r):
            a, b = first_data_row, max(last_block, r - 1)
        else:
            a, b = prev + 1, r - 1
            prev = r
        if b < a:
            a = b = r - 1
        for col, letter in ((5, 'E'), (6, 'F'), (7, 'G')):
            ws_f.cell(r, col).value = f'=SUBTOTAL(9,{letter}{a}:{letter}{b})'
        n += 1
    return n


def capture_row_formulas(ws_f, skip_subtotals: bool = True):
    """Snapshot every formula that is NOT a SUBTOTAL, with its coordinates.

    The CTD carries ~19 of these (the multi-pay pick-up maths at F9/I9, the
    gross-up formulas, the COMP CHECK block, and cross-row references such as
    =H568). openpyxl's insert_rows moves the CELLS but never rewrites the TEXT,
    so each of these must be captured before an insert and rewritten after.
    """
    out = []
    subs = set(subtotal_rows(ws_f)) if skip_subtotals else set()
    for row in ws_f.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith('=') \
                    and 'SUBTOTAL' not in c.value and c.row not in subs:
                out.append((c.row, c.column, c.value))
    return out


def make_remapper(insertions):
    """insertions: [(original_row_position, count), ...] applied in order."""
    def remap(o):
        n, applied = o, 0
        for pos, cnt in insertions:
            if n >= pos + applied:
                n += cnt
            applied += cnt
        return n
    return remap


def restore_row_formulas(ws_f, captured, remap):
    """Rewrite captured formulas at their new coordinates with row refs shifted."""
    def shift(txt):
        return re.sub(r'\b([A-Z]{1,2})(\d+)\b',
                      lambda m: f'{m.group(1)}{remap(int(m.group(2)))}', txt)
    for r, c, txt in captured:
        ws_f.cell(remap(r), c).value = shift(txt)


def copy_row_style(ws, src_row, dst_row, ncols=12):
    for c in range(1, ncols + 1):
        s, d = ws.cell(src_row, c), ws.cell(dst_row, c)
        d.font = copy(s.font); d.border = copy(s.border); d.fill = copy(s.fill)
        d.number_format = s.number_format; d.alignment = copy(s.alignment)


def zero_and_mark(ws, row, reason: str, ncols: int = 9):
    """Neutralise a row without deleting it.

    Deleting shifts every row below and breaks both the SUBTOTAL ranges and the
    cross-row formulas above. Zeroing costs nothing, needs no repair, and leaves
    the audit trail visible to whoever reviews the claim.
    """
    old = str(ws.cell(row, 4).value or '')
    ws.cell(row, 4).value = f'{REMOVED_PREFIX} - {reason}]  {old}'
    for col in (5, 6, 7, 8):
        ws.cell(row, col).value = 0
    for c in range(2, ncols + 1):
        ws.cell(row, c).font = RED
        ws.cell(row, c).fill = GREY


def freeze_external_links(wb, cached: dict | None = None):
    """Replace '[n]OtherWorkbook'!Cell formulas with their cached values.

    openpyxl strips the cached value on save; LibreOffice then cannot resolve the
    link, writes #NAME? and deletes every external reference. recalc.py refuses to
    run in that state. The Everton template carries one at Detail Sheet!N96.
    """
    frozen = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and re.search(r"\[\d+\]", c.value):
                    key = f'{ws.title}!{c.coordinate}'
                    val = (cached or {}).get(key)
                    if val is not None:
                        c.value = val
                        frozen.append((key, val))
    return frozen


def cached_external_values(path):
    """Read the cached values of external-link cells BEFORE editing the workbook."""
    from openpyxl import load_workbook
    wv = load_workbook(path, data_only=True)
    wf = load_workbook(path, data_only=False)
    out = {}
    for ws in wf.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and re.search(r"\[\d+\]", c.value):
                    out[f'{ws.title}!{c.coordinate}'] = wv[ws.title][c.coordinate].value
    return out
